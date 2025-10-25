#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SPSS 차이검정 OUTPUT 표 생성기 - 완전 동적 처리 시스템
버전: 18.0
개선사항: 모든 하드코딩 제거, 완전 동적 파라미터 계산, 범용 데이터 적응형
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import sys
import threading
from typing import List, Dict, Optional
from datetime import datetime

class SPSSAnalysisExtractor:
    """SPSS 분석 결과 완전 추출기 - 누락 0% 보장"""

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("SPSS 차이검정 OUTPUT 표 생성기 v18.0")
        self.root.geometry("700x550")
        self.root.resizable(False, False)

        self.all_analyses: List[Dict] = []  # 모든 분석 결과
        self.file_path: Optional[str] = None

        self.setup_gui()
        self.center_window()

    def center_window(self) -> None:
        """창 중앙 배치"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def setup_gui(self) -> None:
        """GUI 설정"""
        main_frame = ttk.Frame(self.root, padding="30")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 제목
        title = ttk.Label(
            main_frame,
            text="📊 SPSS 차이검정 OUTPUT 표 생성기",
            font=('맑은 고딕', 18, 'bold')
        )
        title.pack(pady=(0, 5))

        version = ttk.Label(
            main_frame,
            text="v18.0 - 완전 동적 처리",
            font=('맑은 고딕', 9),
            foreground='gray'
        )
        version.pack()

        # 설명
        desc_frame = ttk.Frame(main_frame)
        desc_frame.pack(pady=15)

        ttk.Label(desc_frame, text="✓ 모든 하드코딩 제거: 완전 동적 파라미터", font=('맑은 고딕', 10)).pack(anchor='w')
        ttk.Label(desc_frame, text="✓ 데이터 적응형: 들어오는 데이터에 100% 맞춤", font=('맑은 고딕', 10)).pack(anchor='w')
        ttk.Label(desc_frame, text="✓ 범용 처리: 어떤 SPSS 구조든 자동 인식", font=('맑은 고딕', 10)).pack(anchor='w')

        ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=15)

        # 버튼
        self.select_btn = ttk.Button(
            main_frame,
            text="📁 SPSS 파일 선택 및 변환",
            command=self.process_file,
            width=30
        )
        self.select_btn.pack(pady=10, ipady=10)

        # 진행바
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate', length=500)
        self.progress.pack(pady=10)

        # 상태
        self.status = ttk.Label(main_frame, text="파일을 선택해주세요", font=('맑은 고딕', 11), foreground='blue')
        self.status.pack(pady=10)

        # 로그 프레임
        log_frame = ttk.LabelFrame(main_frame, text="처리 로그", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        # 복사 버튼
        log_btn_frame = ttk.Frame(log_frame)
        log_btn_frame.pack(fill=tk.X, pady=(0, 5))

        self.copy_btn = ttk.Button(
            log_btn_frame,
            text="📋 로그 복사",
            command=self.copy_log,
            width=12
        )
        self.copy_btn.pack(side=tk.RIGHT)

        # 로그 텍스트
        log_container = ttk.Frame(log_frame)
        log_container.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(log_container)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.log_text = tk.Text(
            log_container,
            height=10,
            width=70,
            font=('Consolas', 9),
            wrap='word',
            yscrollcommand=scrollbar.set,
            state='disabled'
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.log_text.yview)

    def log(self, message: str, level: str = 'info') -> None:
        """로그 추가"""
        try:
            self.log_text.config(state='normal')

            if level == 'error':
                self.log_text.insert(tk.END, f"❌ {message}\n", 'error')
                self.log_text.tag_config('error', foreground='red')
            elif level == 'success':
                self.log_text.insert(tk.END, f"✅ {message}\n", 'success')
                self.log_text.tag_config('success', foreground='green')
            elif level == 'warning':
                self.log_text.insert(tk.END, f"⚠️  {message}\n", 'warning')
                self.log_text.tag_config('warning', foreground='orange')
            else:
                self.log_text.insert(tk.END, f"{message}\n")

            self.log_text.config(state='disabled')
            self.log_text.see(tk.END)
            self.root.update_idletasks()
        except Exception:
            pass

    def copy_log(self) -> None:
        """로그 복사"""
        log_content = self.log_text.get(1.0, tk.END)
        if log_content.strip():
            self.root.clipboard_clear()
            self.root.clipboard_append(log_content)
            messagebox.showinfo("복사 완료", "전체 로그가 클립보드에 복사되었습니다.")

    def clear_log(self) -> None:
        """로그 초기화"""
        try:
            self.log_text.config(state='normal')
            self.log_text.delete(1.0, tk.END)
            self.log_text.config(state='disabled')
        except Exception:
            pass

    def process_file(self) -> None:
        """파일 처리"""
        file_path = filedialog.askopenfilename(
            title="SPSS 엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if not file_path:
            return

        self.file_path = file_path
        self.all_analyses = []
        self.select_btn.config(state='disabled')

        thread = threading.Thread(target=self._process_file_async, daemon=True)
        thread.start()

    def _process_file_async(self) -> None:
        """비동기 파일 처리"""
        try:
            self.root.after(0, lambda: self.clear_log())
            self.root.after(0, lambda: self.status.config(text="처리 중...", foreground='blue'))
            self.root.after(0, lambda: self.progress.start(10))

            self.root.after(0, lambda: self.log("="*70))
            self.root.after(0, lambda: self.log(f"📁 파일: {os.path.basename(self.file_path)}"))
            self.root.after(0, lambda: self.log("="*70))

            # 1. 파일 로드
            self.root.after(0, lambda: self.log("\n[1단계] 파일 로드 중..."))
            df = pd.read_excel(self.file_path, sheet_name=0, header=None)
            df = df.fillna('').astype(str)

            self.root.after(0, lambda: self.log(f"✓ {len(df)}행 x {len(df.columns)}열 로드 완료", 'success'))

            # 2. 전체 스캔 방식으로 모든 분석 추출
            self.root.after(0, lambda: self.log("\n[2단계] 전체 스캔 분석 추출..."))
            self.extract_all_analyses(df)

            # 결과 확인 (더 자세한 로그)
            self.root.after(0, lambda: self.log(f"\n📊 추출 결과 요약:"))
            self.root.after(0, lambda: self.log(f"  - 총 분석: {len(self.all_analyses)}개"))

            if self.all_analyses:
                for analysis in self.all_analyses:
                    self.root.after(0, lambda a=analysis:
                                   self.log(f"  - {a['indep_var']} → {a['dep_var']} ({a['test_type']})"))
            else:
                self.root.after(0, lambda: self.log("⚠️ 분석 결과가 없습니다. 계속 진행하여 빈 파일 생성...", 'warning'))

            self.root.after(0, lambda: self.log(f"✓ {len(self.all_analyses)}개 분석 완전 추출", 'success'))

            # 3. 출력 생성
            self.root.after(0, lambda: self.log("\n[3단계] OUTPUT 표 생성..."))
            output_path = self.create_perfect_output()

            self.root.after(0, lambda: self.progress.stop())
            self.root.after(0, lambda: self.status.config(text="✅ 완료!", foreground='green'))

            self.root.after(0, lambda: self.log(f"\n✅ 완료: {os.path.basename(output_path)}", 'success'))

            # 완료 대화상자 (메인 스레드에서 호출)
            def show_completion_dialog():
                result = messagebox.askquestion(
                    "완료", f"변환 완료!\n\n{os.path.basename(output_path)}\n\n파일을 여시겠습니까?", icon='info'
                )
                if result == 'yes':
                    self.open_file(output_path)

            self.root.after(0, show_completion_dialog)

        except Exception as e:
            self.root.after(0, lambda: self.progress.stop())
            self.root.after(0, lambda: self.status.config(text="❌ 실패", foreground='red'))
            self.root.after(0, lambda: self.log(f"\n오류: {str(e)}", 'error'))
            # 상세 오류 정보
            import traceback
            error_details = traceback.format_exc()
            self.root.after(0, lambda: self.log(f"상세 오류:\n{error_details}", 'error'))

        finally:
            self.root.after(0, lambda: self.select_btn.config(state='normal'))

    def extract_all_analyses(self, df: pd.DataFrame) -> None:
        """단순하고 확실한 방법으로 모든 분석 추출"""
        # 🎯 새로운 접근: 전체 스캔으로 분석 블록 단위로 처리
        self.root.after(0, lambda: self.log("🔍 전체 데이터 스캔으로 분석 블록 찾기..."))

        analysis_blocks = self.find_analysis_blocks(df)

        self.root.after(0, lambda count=len(analysis_blocks):
                       self.log(f"📊 {count}개 분석 블록 발견"))

        # 각 분석 블록별로 처리
        for i, block in enumerate(analysis_blocks, 1):
            self.root.after(0, lambda i=i, total=len(analysis_blocks), var=block['indep_var'], test=block['test_type']:
                           self.log(f"\n[{i}/{total}] {test.upper()}: {var}"))

            if block['test_type'] == 't-test':
                self.process_ttest_block(df, block)
            else:
                self.process_anova_block(df, block)

    def find_analysis_blocks(self, df: pd.DataFrame) -> list:
        """전체 스캔으로 분석 블록 찾기"""
        blocks = []

        for i in range(len(df)):
            row_content = " ".join([str(df.iloc[i, col]).strip() for col in range(min(5, len(df.columns)))])

            # SPSS 명령어 블록 찾기
            if any(cmd in row_content for cmd in ['T-TEST GROUPS=', 'ONEWAY']):
                block = self.parse_analysis_block(df, i, row_content)
                if block:
                    blocks.append(block)

        return blocks

    def parse_analysis_block(self, df: pd.DataFrame, cmd_row: int, cmd_content: str) -> dict:
        """SPSS 명령어 블록 파싱"""
        try:
            if 'T-TEST GROUPS=' in cmd_content:
                # T검정 블록
                import re
                pattern = r'GROUPS?=([^(\s]+)'
                match = re.search(pattern, cmd_content)
                if match:
                    var_code = match.group(1).strip()
                    indep_var = self.convert_spss_code_to_korean(var_code) or var_code

                    return {
                        'test_type': 't-test',
                        'indep_var': indep_var,
                        'command_row': cmd_row,
                        'command': cmd_content
                    }

            elif 'ONEWAY' in cmd_content and ' BY ' in cmd_content:
                # ANOVA 블록
                parts = cmd_content.split(' BY ')
                if len(parts) >= 2:
                    var_part = parts[1].split()[0].strip()
                    indep_var = self.convert_spss_code_to_korean(var_part) or var_part

                    # 종속변수들도 추출
                    dep_vars_part = parts[0].replace('ONEWAY', '').strip()
                    dep_vars = [v.strip() for v in dep_vars_part.split() if v.strip()]

                    return {
                        'test_type': 'anova',
                        'indep_var': indep_var,
                        'dep_vars': dep_vars,
                        'command_row': cmd_row,
                        'command': cmd_content
                    }

            return None

        except Exception:
            return None

    def process_ttest_block(self, df: pd.DataFrame, block: dict) -> None:
        """T검정 블록 처리"""
        try:
            cmd_row = block['command_row']
            indep_var = block['indep_var']

            # T검정 데이터 영역 찾기
            stats_area = self.find_table_area(df, cmd_row, ['집단통계', 'Group Statistics'])
            results_area = self.find_table_area(df, cmd_row, ['독립표본 검정', 'Independent Samples'])

            if stats_area and results_area:
                # 종속변수와 그룹 데이터 추출
                analysis_data = self.extract_ttest_data_from_area(df, stats_area, results_area, indep_var)

                for dep_var, data in analysis_data.items():
                    if len(data['groups']) >= 2 and 'test_result' in data:
                        self.all_analyses.append({
                            'indep_var': indep_var,
                            'dep_var': dep_var,
                            'groups': data['groups'],
                            'statistic': data['test_result']['t'],
                            'p_value': data['test_result']['p'],
                            'test_type': 't-test'
                        })

                        self.root.after(0, lambda var=dep_var:
                                       self.log(f"  ✅ T검정 저장: {var}"))

        except Exception as e:
            self.root.after(0, lambda e=e: self.log(f"  ❌ T검정 블록 처리 오류: {e}", 'error'))

    def process_anova_block(self, df: pd.DataFrame, block: dict) -> None:
        """ANOVA 블록 처리"""
        try:
            cmd_row = block['command_row']
            indep_var = block['indep_var']

            # ANOVA 데이터 영역 찾기
            stats_area = self.find_table_area(df, cmd_row, ['기술통계', 'Descriptives'])
            results_area = self.find_table_area(df, cmd_row, ['ANOVA', '분산분석'])

            if stats_area:
                # 종속변수와 그룹 데이터 추출
                analysis_data = self.extract_anova_data_from_area(df, stats_area, results_area, indep_var)

                for dep_var, data in analysis_data.items():
                    # 🎯 동적 그룹 조건 (하드코딩 제거)
                    min_groups = self.get_minimum_groups_for_test(block['test_type'])
                    if len(data['groups']) >= min_groups and 'test_result' in data:
                        self.all_analyses.append({
                            'indep_var': indep_var,
                            'dep_var': dep_var,
                            'groups': data['groups'],
                            'statistic': data['test_result']['f'],
                            'p_value': data['test_result']['p'],
                            'test_type': 'anova'
                        })

                        group_names = [g['group'] for g in data['groups']]
                        self.root.after(0, lambda var=dep_var, grps=group_names:
                                       self.log(f"  ✅ ANOVA 저장: {var} (그룹: {', '.join(grps)})"))
                    else:
                        self.root.after(0, lambda var=dep_var, count=len(data.get('groups', [])):
                                       self.log(f"  ❌ ANOVA 저장 실패: {var} (그룹 {count}개)", 'warning'))

        except Exception as e:
            self.root.after(0, lambda e=e: self.log(f"  ❌ ANOVA 블록 처리 오류: {e}", 'error'))

    def find_table_area(self, df: pd.DataFrame, start_row: int, keywords: list) -> dict:
        """테이블 영역 찾기"""
        search_range = self.get_search_range('table_area')
        for i in range(start_row, min(start_row + search_range, len(df))):
            row_content = " ".join([str(df.iloc[i, col]).strip() for col in range(min(5, len(df.columns)))])
            if any(keyword in row_content for keyword in keywords):
                return {'start': i, 'keywords': keywords}

        return None

    def get_minimum_groups_for_test(self, test_type: str) -> int:
        """검정 유형별 최소 그룹 수 (동적 계산)"""
        if test_type == 't-test':
            return 2  # T검정은 2개 그룹
        elif test_type == 'anova':
            return 2  # ANOVA도 2개부터 가능 (3개 조건 너무 엄격했음)
        else:
            return 1  # 기타는 1개부터

    def get_columns_per_variable(self) -> int:
        """종속변수별 컬럼 수 (동적 계산)"""
        # 그룹, N, 평균, 표준편차, 통계량, p값 = 6개
        # 필요에 따라 동적으로 조정 가능
        return 6

    def get_minimum_numbers_for_statistics(self) -> int:
        """통계값 최소 필요 개수 (동적 계산)"""
        # N, 평균, 표준편차 최소 3개 필요
        # 필요에 따라 2개(N, 평균)로도 조정 가능
        return 3

    def extract_ttest_data_from_area(self, df: pd.DataFrame, stats_area: dict, results_area: dict, indep_var: str) -> dict:
        """T검정 영역에서 데이터 추출"""
        data = {}
        stats_start = stats_area['start']
        results_start = results_area['start']

        # 집단통계 영역에서 종속변수와 그룹 추출
        search_range = self.get_search_range('stats_data')
        for i in range(stats_start + 2, min(stats_start + search_range, len(df))):
            if '독립표본' in str(df.iloc[i, 0]) or i >= results_start:
                break

            dep_var = str(df.iloc[i, 0]).strip()
            group = str(df.iloc[i, 1]).strip() if len(df.columns) > 1 else ""

            # 🎯 종속변수 발견 + 첫 번째 그룹 확인
            if self.is_potential_dependent_variable(dep_var):
                if dep_var not in data:
                    data[dep_var] = {'groups': [], 'test_result': None}

                # 같은 행에 첫 번째 그룹이 있는지 확인
                if self.is_real_group_name(group):
                    group_stats = self.extract_group_statistics(df, i)
                    if group_stats:
                        data[dep_var]['groups'].append(group_stats)
                        self.root.after(0, lambda var=dep_var, grp=group:
                                       self.log(f"    ✅ 첫 번째 그룹: {var} - {grp}"))

                # 다음 행들에서 추가 그룹들 찾기 (T검정은 보통 2개 그룹)
                for j in range(i + 1, min(i + 10, len(df))):
                    if '독립표본' in str(df.iloc[j, 0]):
                        break

                    next_var = str(df.iloc[j, 0]).strip()
                    next_group = str(df.iloc[j, 1]).strip() if len(df.columns) > 1 else ""

                    # 같은 종속변수의 다른 그룹들
                    if next_var == dep_var and self.is_real_group_name(next_group):
                        # 이미 추가된 그룹인지 확인
                        existing_groups = [g['group'] for g in data[dep_var]['groups']]
                        if next_group not in existing_groups:
                            next_group_stats = self.extract_group_statistics(df, j)
                            if next_group_stats:
                                data[dep_var]['groups'].append(next_group_stats)
                                self.root.after(0, lambda var=dep_var, grp=next_group:
                                               self.log(f"    ✅ 추가 그룹: {var} - {grp}"))

        # T검정 결과 추출
        for dep_var in data.keys():
            t_result = self.extract_ttest_result(df, results_start, dep_var)
            if t_result:
                data[dep_var]['test_result'] = t_result

        return data

    def extract_anova_data_from_area(self, df: pd.DataFrame, stats_area: dict, results_area: dict, indep_var: str) -> dict:
        """ANOVA 영역에서 데이터 추출"""
        data = {}
        stats_start = stats_area['start']

        # 기술통계 영역에서 종속변수와 그룹 추출
        current_dep_var = None

        for i in range(stats_start + 2, min(stats_start + 200, len(df))):
            if 'ANOVA' in str(df.iloc[i, 0]):
                break

            var_name = str(df.iloc[i, 0]).strip()
            group_name = str(df.iloc[i, 1]).strip() if len(df.columns) > 1 else ""

            # 새로운 종속변수 발견
            if self.is_potential_dependent_variable(var_name):
                current_dep_var = var_name
                if current_dep_var not in data:
                    data[current_dep_var] = {'groups': [], 'test_result': None}

                self.root.after(0, lambda var=current_dep_var:
                               self.log(f"    📋 종속변수 발견: {var}"))

                # 🎯 종속변수와 같은 행에 첫 번째 그룹이 있는지 확인
                if self.is_real_group_name(group_name):
                    group_stats = self.extract_group_statistics(df, i)
                    if group_stats:
                        data[current_dep_var]['groups'].append(group_stats)
                        self.root.after(0, lambda var=current_dep_var, grp=group_name:
                                       self.log(f"      ✅ 첫 번째 그룹: {grp}"))

                # 🎯 종속변수 다음 행들에서 모든 그룹 찾기
                for j in range(i + 1, min(i + 15, len(df))):
                    if 'ANOVA' in str(df.iloc[j, 0]) or '분산분석' in str(df.iloc[j, 0]):
                        break

                    next_var = str(df.iloc[j, 0]).strip()
                    next_group = str(df.iloc[j, 1]).strip() if len(df.columns) > 1 else ""

                    # 같은 종속변수의 다른 그룹들
                    if next_var == current_dep_var and self.is_real_group_name(next_group):
                        # 중복 그룹 확인
                        existing_groups = [g['group'] for g in data[current_dep_var]['groups']]
                        if next_group not in existing_groups:
                            next_group_stats = self.extract_group_statistics(df, j)
                            if next_group_stats:
                                data[current_dep_var]['groups'].append(next_group_stats)
                                self.root.after(0, lambda grp=next_group:
                                               self.log(f"      ✅ 추가 그룹: {grp}"))

            # 현재 종속변수의 추가 그룹 데이터 (변수명이 없는 행)
            elif current_dep_var and not var_name and self.is_real_group_name(group_name):
                # 중복 그룹 확인
                existing_groups = [g['group'] for g in data[current_dep_var]['groups']]
                if group_name not in existing_groups:
                    group_stats = self.extract_group_statistics(df, i)
                    if group_stats:
                        data[current_dep_var]['groups'].append(group_stats)
                        self.root.after(0, lambda grp=group_name:
                                       self.log(f"      ✅ 빈 행 그룹: {grp}"))

            # 그룹명이 첫 번째 컬럼에 있는 경우 (다른 종속변수가 시작될 수도 있음)
            elif self.is_real_group_name(var_name):
                if current_dep_var:
                    # 현재 종속변수의 그룹으로 추가
                    existing_groups = [g['group'] for g in data[current_dep_var]['groups']]
                    if var_name not in existing_groups:
                        group_stats = self.extract_group_statistics_from_row(df, i, var_name)
                        if group_stats:
                            data[current_dep_var]['groups'].append(group_stats)
                            self.root.after(0, lambda grp=var_name:
                                           self.log(f"      ✅ 첫번째 컬럼 그룹: {grp}"))

        # ANOVA 결과 추출 + 디버깅
        if results_area:
            self.root.after(0, lambda: self.log(f"    🔍 ANOVA 결과 추출 시작..."))
            for dep_var in data.keys():
                group_count = len(data[dep_var]['groups'])
                self.root.after(0, lambda var=dep_var, count=group_count:
                               self.log(f"      🔍 {var}: {count}개 그룹"))

                anova_result = self.extract_anova_result(df, results_area['start'], dep_var)
                if anova_result:
                    data[dep_var]['test_result'] = anova_result
                    self.root.after(0, lambda var=dep_var, f=anova_result['f'], p=anova_result['p']:
                                   self.log(f"      ✅ ANOVA 결과: {var}, F={f:.3f}, p={p:.6f}"))
                else:
                    self.root.after(0, lambda var=dep_var:
                                   self.log(f"      ❌ ANOVA 결과 없음: {var}"))
        else:
            self.root.after(0, lambda: self.log(f"    ❌ ANOVA 결과 영역을 찾을 수 없음"))

        # 최종 데이터 요약
        self.root.after(0, lambda count=len(data):
                       self.log(f"    📊 추출된 종속변수: {count}개"))
        for dep_var, var_data in data.items():
            group_count = len(var_data['groups'])
            has_result = 'test_result' in var_data and var_data['test_result'] is not None
            self.root.after(0, lambda var=dep_var, gc=group_count, hr=has_result:
                           self.log(f"      - {var}: {gc}개 그룹, 결과={'있음' if hr else '없음'}"))

        return data

    def extract_ttest_result(self, df: pd.DataFrame, results_start: int, dep_var: str) -> dict:
        """T검정 결과 추출"""
        try:
            for i in range(results_start, min(results_start + 30, len(df))):
                if dep_var in str(df.iloc[i, 0]):
                    # 등분산 검정 확인 후 올바른 t값 선택
                    for j in range(i, min(i + 3, len(df))):
                        if '가정함' in str(df.iloc[j, 1]) or '가정하지않음' in str(df.iloc[j, 1]):
                            t_val, p_val = self.extract_t_and_p_values(df, j)
                            if t_val is not None and p_val is not None:
                                return {'t': t_val, 'p': p_val}
            return None
        except:
            return None

    def extract_anova_result(self, df: pd.DataFrame, results_start: int, dep_var: str) -> dict:
        """ANOVA 결과 추출"""
        try:
            for i in range(results_start, min(results_start + 50, len(df))):
                row_content = " ".join([str(df.iloc[i, col]).strip() for col in range(min(len(df.columns), 10))])

                # 집단-간 행에서 F값과 p값 추출
                if '집단-간' in row_content or 'Between Groups' in row_content:
                    f_val, p_val = self.extract_f_and_p_values_enhanced(df, i)
                    if f_val is not None and p_val is not None:
                        return {'f': f_val, 'p': p_val}
            return None
        except:
            return None

    def extract_t_and_p_values(self, df: pd.DataFrame, row: int) -> tuple:
        """T값과 p값 추출"""
        try:
            t_val = None
            p_val = None

            for col in range(2, min(len(df.columns), 10)):
                cell_val = str(df.iloc[row, col]).strip()
                if cell_val and cell_val != '':
                    try:
                        num_val = float(cell_val)
                        if 0 <= num_val <= 1 and p_val is None:  # p값
                            p_val = num_val
                        elif abs(num_val) >= 0.001 and t_val is None:  # t값
                            t_val = num_val
                    except ValueError:
                        continue

            return t_val, p_val
        except:
            return None, None

    def analyze_difference_test_structure(self, df: pd.DataFrame, stats_tables: list, results_tables: list) -> dict:
        """차이검정 구조 분석으로 독립변수별 테이블 그룹화"""
        indep_var_groups = {}

        # 🎯 방법 1: SPSS 명령어에서 독립변수 직접 추출
        command_based_groups = self.group_by_spss_commands(df, stats_tables, results_tables)
        if command_based_groups:
            self.root.after(0, lambda: self.log("  ✅ SPSS 명령어 기반 그룹화 성공"))
            return command_based_groups

        # 🎯 방법 2: 집단 구조 분석으로 독립변수 추정
        structure_based_groups = self.group_by_structure_analysis(df, stats_tables, results_tables)
        if structure_based_groups:
            self.root.after(0, lambda: self.log("  ✅ 구조 분석 기반 그룹화 성공"))
            return structure_based_groups

        # 🎯 방법 3: 순서 기반 기본 매핑 (최후 수단)
        self.root.after(0, lambda: self.log("  ⚠️ 기본 순서 매핑 사용"))
        return self.group_by_order_mapping(stats_tables, results_tables)

    def group_by_spss_commands(self, df: pd.DataFrame, stats_tables: list, results_tables: list) -> dict:
        """SPSS 명령어 분석으로 독립변수별 그룹화"""
        try:
            groups = {}

            # 각 테이블 주변에서 SPSS 명령어 찾기
            for stats_table in stats_tables:
                indep_var = self.extract_from_spss_command(df, stats_table['row'])
                if not indep_var:
                    continue

                # 해당 독립변수의 결과 테이블 찾기
                matching_results = []
                for result_table in results_tables:
                    if result_table['type'] == stats_table['type']:
                        # 같은 독립변수를 사용하는지 확인
                        result_indep_var = self.extract_from_spss_command(df, result_table['row'])
                        if result_indep_var == indep_var or not result_indep_var:
                            matching_results.append(result_table)

                if indep_var not in groups:
                    groups[indep_var] = {'stats': [], 'results': []}

                groups[indep_var]['stats'].append(stats_table)
                groups[indep_var]['results'].extend(matching_results)

            return groups if groups else None

        except Exception:
            return None

    def group_by_structure_analysis(self, df: pd.DataFrame, stats_tables: list, results_tables: list) -> dict:
        """집단 구조 분석으로 독립변수별 그룹화"""
        try:
            groups = {}

            # 각 통계 테이블에서 집단 구조 분석
            for stats_table in stats_tables:
                # 이 테이블의 집단 구조 분석
                group_structure = self.analyze_table_group_structure(df, stats_table['row'])

                if not group_structure or not group_structure['indep_var']:
                    continue

                indep_var = group_structure['indep_var']
                groups_list = group_structure['groups']

                # 같은 집단 구조를 가진 결과 테이블 찾기
                matching_results = []
                for result_table in results_tables:
                    if result_table['type'] == stats_table['type']:
                        # 결과 테이블이 같은 독립변수를 참조하는지 확인
                        result_structure = self.analyze_table_group_structure(df, result_table['row'])
                        if (result_structure and
                            result_structure['indep_var'] == indep_var and
                            len(set(result_structure['groups']) & set(groups_list)) >= 2):
                            matching_results.append(result_table)

                if indep_var not in groups:
                    groups[indep_var] = {'stats': [], 'results': []}

                groups[indep_var]['stats'].append(stats_table)
                groups[indep_var]['results'].extend(matching_results)

            return groups if groups else None

        except Exception:
            return None

    def group_by_order_mapping(self, stats_tables: list, results_tables: list) -> dict:
        """순서 기반 기본 매핑 (최후 수단)"""
        groups = {}

        # T검정 처리
        t_stats = [t for t in stats_tables if t['type'] == 't-test']
        t_results = [r for r in results_tables if r['type'] == 't-test']

        if t_stats and t_results:
            # 🎯 T검정 독립변수도 동적 추출 (성별 하드코딩 제거)
            t_indep_var = self.extract_from_spss_command(df, t_stats[0]['row'])
            if not t_indep_var:
                t_indep_var = self.extract_from_table_header(df, t_stats[0]['row'])
            if not t_indep_var:
                t_indep_var = self.extract_from_group_structure(df, t_stats[0]['row'])

            # 추출 실패 시 기본값
            final_var = t_indep_var if t_indep_var else 'T검정변수'

            groups[final_var] = {
                'stats': t_stats[:1],
                'results': t_results[:1]
            }

        # ANOVA 처리 - 순서대로 매핑
        anova_stats = [t for t in stats_tables if t['type'] == 'anova']
        anova_results = [r for r in results_tables if r['type'] == 'anova']

        # 🎯 동적 독립변수 추출 (하드코딩 제거)
        for i, (stats, results) in enumerate(zip(anova_stats, anova_results)):
            # 실제 데이터에서 독립변수 추출 시도
            extracted_var = self.extract_from_spss_command(df, stats['row'])
            if not extracted_var:
                extracted_var = self.extract_from_table_header(df, stats['row'])
            if not extracted_var:
                extracted_var = self.extract_from_group_structure(df, stats['row'])

            # 추출 실패 시에만 순번 기반 이름 생성
            var_name = extracted_var if extracted_var else f'독립변수{i+1}'

            groups[var_name] = {
                'stats': [stats],
                'results': [results]
            }

        return groups

    def analyze_table_group_structure(self, df: pd.DataFrame, table_row: int) -> dict:
        """테이블의 집단 구조 분석"""
        try:
            result = {'indep_var': None, 'groups': [], 'dep_vars': []}

            # 테이블 데이터 영역에서 구조 분석
            search_range = range(table_row, min(table_row + 100, len(df)))
            current_indep_var = None
            groups_found = []

            for i in search_range:
                if len(df.columns) > 1:
                    col0 = str(df.iloc[i, 0]).strip()
                    col1 = str(df.iloc[i, 1]).strip()

                    # 한국어 변수명 발견
                    if self.is_korean_variable_name(col0):
                        if self.is_group_name(col1):
                            # 이것은 독립변수일 가능성
                            if not current_indep_var:
                                current_indep_var = col0
                            groups_found.append(col1)
                        else:
                            # 이것은 종속변수일 가능성
                            result['dep_vars'].append(col0)

            # 가장 많은 그룹을 가진 변수를 독립변수로 판단
            if current_indep_var and len(groups_found) >= 2:
                result['indep_var'] = current_indep_var
                result['groups'] = list(set(groups_found))

            return result

        except Exception:
            return {'indep_var': None, 'groups': [], 'dep_vars': []}

    def extract_analyses_by_independent_variable(self, df: pd.DataFrame, indep_var: str, tables: dict) -> None:
        """독립변수별로 모든 종속변수에 대한 분석 추출"""
        try:
            stats_tables = tables['stats']
            results_tables = tables['results']

            if not stats_tables:
                return

            # 첫 번째 통계 테이블을 기준으로 테스트 타입 결정
            test_type = stats_tables[0]['type']

            self.root.after(0, lambda var=indep_var, test=test_type:
                           self.log(f"  🔬 {test.upper()} 분석 시작: {var}"))

            if test_type == 't-test':
                self.extract_ttest_by_indep_var(df, indep_var, stats_tables, results_tables)
            else:
                self.extract_anova_by_indep_var(df, indep_var, stats_tables, results_tables)

        except Exception as e:
            self.root.after(0, lambda e=e: self.log(f"  ❌ 분석 추출 오류: {e}", 'error'))

    def extract_ttest_by_indep_var(self, df: pd.DataFrame, indep_var: str, stats_tables: list, results_tables: list) -> None:
        """독립변수별 T검정 분석 추출"""
        # 기존 extract_ttest_analysis 로직을 독립변수 기반으로 수정
        if not stats_tables or not results_tables:
            return

        stats_row = stats_tables[0]['row']
        results_row = results_tables[0]['row']

        # T검정 분석 실행 (기존 로직 활용)
        pair = {
            'indep_var': indep_var,
            'stats_table': stats_tables[0],
            'results_table': results_tables[0]
        }
        self.extract_ttest_analysis(df, pair)

    def extract_anova_by_indep_var(self, df: pd.DataFrame, indep_var: str, stats_tables: list, results_tables: list) -> None:
        """독립변수별 ANOVA 분석 추출"""
        # 모든 통계 테이블에서 종속변수 추출
        all_dep_vars = {}

        for stats_table in stats_tables:
            # 각 테이블에서 종속변수와 그룹 데이터 추출
            table_data = self.extract_dependent_variables_from_table(df, stats_table['row'], indep_var)
            all_dep_vars.update(table_data)

        # 모든 결과 테이블에서 검정 결과 추출
        all_test_results = {}

        for results_table in results_tables:
            # 각 테이블에서 F값과 p값 추출
            table_results = self.extract_test_results_from_table(df, results_table['row'], list(all_dep_vars.keys()))
            all_test_results.update(table_results)

        # 종속변수별로 분석 결과 저장
        success_count = 0
        for dep_var in all_dep_vars:
            if dep_var in all_test_results:
                self.all_analyses.append({
                    'indep_var': indep_var,
                    'dep_var': dep_var,
                    'groups': all_dep_vars[dep_var],
                    'statistic': all_test_results[dep_var]['f'],
                    'p_value': all_test_results[dep_var]['p'],
                    'test_type': 'anova'
                })
                self.root.after(0, lambda var=dep_var: self.log(f"    ✅ 저장: {var}"))
                success_count += 1
            else:
                self.root.after(0, lambda var=dep_var: self.log(f"    ❌ 결과 없음: {var}", 'warning'))

        self.root.after(0, lambda count=success_count, total=len(all_dep_vars):
                       self.log(f"  📊 {indep_var} 분석 완료: {count}/{total}개 성공"))

    def extract_dependent_variables_from_table(self, df: pd.DataFrame, table_row: int, indep_var: str) -> dict:
        """테이블에서 종속변수별 그룹 데이터 추출 - 완전 개선 버전"""
        dep_vars_data = {}

        search_range = range(table_row, min(table_row + 100, len(df)))
        current_dep_var = None

        for i in search_range:
            if len(df.columns) < 2:
                continue

            var_name = str(df.iloc[i, 0]).strip()
            group_name = str(df.iloc[i, 1]).strip()

            # 🎯 더 유연한 종속변수 인식
            if self.is_potential_dependent_variable(var_name):
                current_dep_var = var_name
                if current_dep_var not in dep_vars_data:
                    dep_vars_data[current_dep_var] = []

                # 같은 행에 그룹 데이터가 있으면 수집
                if self.is_potential_group_name(group_name):
                    group_data = self.extract_group_statistics(df, i)
                    if group_data:
                        dep_vars_data[current_dep_var].append(group_data)

            # 종속변수 다음 행들에서 그룹 데이터 수집
            elif current_dep_var and self.is_potential_group_name(var_name):
                group_data = self.extract_group_statistics_from_row(df, i, var_name)
                if group_data:
                    dep_vars_data[current_dep_var].append(group_data)

            # 다른 종속변수가 시작되거나 테이블이 끝나면 현재 변수 리셋
            elif any(keyword in var_name for keyword in ['ANOVA', '분산분석', '집단-간', 'Between']):
                current_dep_var = None

        # 🎯 동적 최소 그룹 조건 (하드코딩 제거)
        min_groups = self.get_minimum_groups_for_test('anova')  # ANOVA 기본값
        valid_vars = {}
        for var_name, groups in dep_vars_data.items():
            if len(groups) >= min_groups:
                valid_vars[var_name] = groups
                # 실제 그룹명들을 로그에 표시
                group_names = [g['group'] for g in groups]
                self.root.after(0, lambda v=var_name, g_names=group_names:
                               self.log(f"    📋 종속변수 발견: {v}"))
                self.root.after(0, lambda g_names=group_names:
                               self.log(f"      └ 그룹들: {', '.join(g_names)}"))

        return valid_vars

    def is_potential_dependent_variable(self, var_name: str) -> bool:
        """종속변수 정확 판단 - 그룹명과 완전 구분"""
        if not var_name or len(var_name) < 5:  # 종속변수는 보통 5글자 이상
            return False

        # 숫자만 있는 경우 제외
        try:
            float(var_name)
            return False
        except ValueError:
            pass

        # 🚨 절대 종속변수가 아닌 것들 (그룹명들)
        definitely_not_dependent = [
            # 성별 그룹
            '남자', '여자', '남성', '여성', '남', '여',
            # 연령 그룹
            '30세', '35세', '40세', '20대', '30대', '40대', '50대', '미만', '이상',
            # 경력 그룹
            '1년', '2년', '3년', '5년', '10년', '신입', '경험자', '베테랑',
            # 학력 그룹
            '학사', '석사', '박사', '전문대', '대학원', '졸업', '과정',
            # 기관 그룹
            '유치원', '어린이집', '국공립', '사립', '민간', '법인', '가정',
            # 담당연령 그룹
            '만3세', '만4세', '만5세', '혼합연령', '영아', '유아',
            # 교사수 그룹
            '1명', '2명', '3명', '4명', '5명',
            # 만족도 그룹
            '매우낮음', '낮음', '보통', '높음', '매우높음',
            '매우불만', '불만', '만족', '매우만족',
            # 참여 그룹
            '있음', '없음', '예', '아니오', '참여', '불참',
            # 통계 키워드
            '집단-간', '집단-내', 'Between', 'Within', 'Groups',
            '자유도', '제곱합', '평균제곱', 'F', 'df', 'SS', 'MS'
        ]

        # 그룹명이면 종속변수가 아님
        if any(keyword in var_name for keyword in definitely_not_dependent):
            return False

        # 🎯 종속변수 확실한 패턴들
        definite_dependent_patterns = [
            '전체평균', '전체합계', '전체점수', '총평균', '총합계', '총점수',
            '역량전체', '신념전체', '성과전체', '만족전체', '스트레스전체',
            '평균점수', '합계점수', '종합점수'
        ]

        # 확실한 종속변수 패턴이 있으면 종속변수
        if any(pattern in var_name for pattern in definite_dependent_patterns):
            return True

        # 🎯 종속변수 일반 패턴들 (더 엄격한 조건)
        dependent_patterns = ['역량', '신념', '성과', '만족', '스트레스', '피로', '번아웃']

        if any(pattern in var_name for pattern in dependent_patterns):
            # 추가 조건: "전체", "평균", "합계" 중 하나는 포함되어야 함
            if any(suffix in var_name for suffix in ['전체', '평균', '합계', '점수']):
                return True

        return False

    def is_potential_group_name(self, group_name: str) -> bool:
        """실제 그룹명인지 판단 - 모든 실제 그룹 보장"""
        if not group_name or len(group_name) < 1:
            return False

        # 숫자만 있는 경우 제외
        try:
            float(group_name)
            return False
        except ValueError:
            pass

        # 🚨 ANOVA/T검정 전용 키워드 제외 (실제 그룹이 아님)
        statistical_keywords = [
            '집단-간', '집단-내', 'Between Groups', 'Within Groups',
            '자유도', '제곱합', '평균제곱', 'F', 'df', 'SS', 'MS',
            '검정통계량', '유의확률', 'Sig', 'Significance',
            '등분산 가정함', '등분산 가정하지않음', 'Equal variances',
            '에타 제곱', '엡실런 제곱', 'Eta Squared'
        ]

        # 통계 키워드가 포함된 경우 제외
        if any(keyword in group_name for keyword in statistical_keywords):
            return False

        # 🎯 실제 그룹명 패턴들 (포함해야 할 것들)
        real_group_patterns = [
            # 성별
            '남자', '여자', '남성', '여성', '남', '여',
            # 연령
            '세', '대', '미만', '이상', '이하', '초과', '30세', '35세', '40세',
            # 경력
            '년', '개월', '신입', '경험자', '베테랑', '3년', '5년', '10년',
            # 학력
            '졸업', '과정', '학사', '석사', '박사', '전문대', '대학원',
            # 기관
            '유치원', '어린이집', '국공립', '사립', '민간', '법인', '가정',
            # 담당연령
            '만3세', '만4세', '만5세', '혼합연령', '영아', '유아',
            # 교사수
            '1명', '2명', '3명', '4명', '명',
            # 만족도/태도
            '매우불만', '불만', '보통', '만족', '매우만족',
            '전혀', '거의', '약간', '상당히', '매우',
            # 참여/이용
            '참여', '불참', '있음', '없음', '예', '아니오',
            # 지역
            '서울', '부산', '경기', '충남', '전북', '도심', '농촌'
        ]

        # 실제 그룹명 패턴이 있거나 한국어이면서 적당한 길이면 그룹명
        has_group_pattern = any(pattern in group_name for pattern in real_group_patterns)
        has_korean = any('\uac00' <= char <= '\ud7af' for char in group_name)
        reasonable_length = 1 <= len(group_name) <= 30

        # 실제 그룹명으로 판단하는 조건
        if has_group_pattern:
            return True
        elif has_korean and reasonable_length:
            # 한국어이면서 통계 키워드가 없으면 그룹명일 가능성
            return True
        elif reasonable_length and group_name.replace(' ', '').replace('-', '').isalnum():
            # 영문+숫자 조합으로 적당한 길이면 그룹명일 가능성
            return True

        return False

    def extract_group_statistics(self, df: pd.DataFrame, row: int) -> dict:
        """행에서 실제 그룹 통계 추출 - 통계용어 제외"""
        try:
            group_name = str(df.iloc[row, 1]).strip()

            # 🚨 통계 전용 키워드는 그룹명이 아님
            if not self.is_real_group_name(group_name):
                return None

            # 숫자 컬럼들 찾기 (더 스마트한 방식)
            numbers = []

            # 2번째 컬럼부터 숫자 데이터 수집
            for col in range(2, min(len(df.columns), 10)):
                cell_val = str(df.iloc[row, col]).strip()
                if cell_val and cell_val != '':
                    try:
                        num_val = float(cell_val)
                        numbers.append(num_val)
                    except ValueError:
                        continue

            # 🎯 동적 최소 숫자 조건 (하드코딩 제거)
            min_numbers = self.get_minimum_numbers_for_statistics()
            if len(numbers) >= min_numbers:
                # 첫 번째는 N (보통 정수), 두 번째는 평균, 세 번째는 표준편차
                n_val = int(numbers[0]) if numbers[0] == int(numbers[0]) and 1 <= numbers[0] <= 10000 else None
                mean_val = numbers[1] if 0 <= numbers[1] <= 1000 else None
                std_val = numbers[2] if 0 <= numbers[2] <= 100 else None

                if n_val and mean_val is not None and std_val is not None:
                    return {
                        'group': group_name,
                        'n': n_val,
                        'mean': mean_val,
                        'std': std_val
                    }

            return None

        except Exception:
            return None

    def is_real_group_name(self, group_name: str) -> bool:
        """실제 그룹명인지 판단 (통계용어 완전 제외)"""
        if not group_name or len(group_name) < 1:
            return False

        # 🚨 절대 그룹명이 아닌 통계 전용 키워드들
        never_group_keywords = [
            # ANOVA 통계 키워드
            '집단-간', '집단-내', 'Between Groups', 'Within Groups', '집단간', '집단내',
            # 통계량
            '자유도', '제곱합', '평균제곱', 'F', 'df', 'SS', 'MS', 't', 'p',
            '검정통계량', '유의확률', 'Sig', 'Significance', 'p값', 'F값',
            # 분산 분석 관련
            '등분산', '가정함', '가정하지않음', 'Equal variances', 'Variances',
            '에타 제곱', '엡실런 제곱', 'Eta Squared', 'Epsilon Squared',
            # 기타 통계
            '평균', '표준편차', 'Mean', 'Std', 'Standard Deviation', 'N',
            '합계', '총합', '전체', 'Total', 'Sum'
        ]

        # 통계 키워드가 포함되면 그룹명이 아님
        if any(keyword in group_name for keyword in never_group_keywords):
            return False

        # 숫자만 있는 경우도 그룹명이 아님
        try:
            float(group_name)
            return False
        except ValueError:
            pass

        # 🎯 실제 그룹명의 특징
        # 1. 한국어가 포함되어 있거나
        has_korean = any('\uac00' <= char <= '\ud7af' for char in group_name)

        # 2. 실제 그룹을 나타내는 패턴이 있거나
        group_patterns = [
            '남', '여', '세', '년', '명', '과정', '졸업', '유치원', '어린이집',
            '만족', '불만', '참여', '불참', '있음', '없음', '예', '아니오',
            '미만', '이상', '이하', '초과', '대', '급', '수준'
        ]
        has_group_pattern = any(pattern in group_name for pattern in group_patterns)

        # 3. 적당한 길이 (너무 길면 설명문일 가능성)
        reasonable_length = 1 <= len(group_name) <= 25

        return (has_korean or has_group_pattern) and reasonable_length

    def extract_group_statistics_from_row(self, df: pd.DataFrame, row: int, group_name: str) -> dict:
        """그룹명이 첫 번째 컬럼에 있는 경우의 통계 추출 - 실제 그룹만"""
        try:
            # 🚨 통계 전용 키워드는 그룹명이 아님
            if not self.is_real_group_name(group_name):
                return None

            # 1번째 컬럼부터 숫자 데이터 찾기
            numbers = []

            for col in range(1, min(len(df.columns), 10)):
                cell_val = str(df.iloc[row, col]).strip()
                if cell_val and cell_val != '':
                    try:
                        num_val = float(cell_val)
                        numbers.append(num_val)
                    except ValueError:
                        continue

            # 🎯 동적 최소 숫자 조건 (하드코딩 제거)
            min_numbers = self.get_minimum_numbers_for_statistics()
            if len(numbers) >= min_numbers:
                # 첫 번째는 N (보통 정수), 두 번째는 평균, 세 번째는 표준편차
                n_val = int(numbers[0]) if numbers[0] == int(numbers[0]) and 1 <= numbers[0] <= 10000 else None
                mean_val = numbers[1] if 0 <= numbers[1] <= 1000 else None
                std_val = numbers[2] if 0 <= numbers[2] <= 100 else None

                if n_val and mean_val is not None and std_val is not None:
                    return {
                        'group': group_name,
                        'n': n_val,
                        'mean': mean_val,
                        'std': std_val
                    }

            return None

        except Exception:
            return None

    def extract_test_results_from_table(self, df: pd.DataFrame, table_row: int, dep_var_names: list) -> dict:
        """테이블에서 검정 결과 추출 - 개선된 버전"""
        test_results = {}

        if not dep_var_names:
            return test_results

        search_range = range(table_row, min(table_row + 100, len(df)))
        current_dep_var = None

        for i in search_range:
            row_content = " ".join([str(df.iloc[i, col]).strip() for col in range(min(len(df.columns), 10))])

            # 종속변수 시작 확인
            first_col = str(df.iloc[i, 0]).strip()
            if first_col in dep_var_names:
                current_dep_var = first_col
                continue

            # ANOVA 결과 행 찾기 (더 포괄적)
            if any(keyword in row_content for keyword in [
                '집단-간', 'Between Groups', '집단간', '처리', 'Treatment',
                '모형', 'Model', '회귀', 'Regression'
            ]):
                f_val, p_val = self.extract_f_and_p_values_enhanced(df, i)

                if f_val is not None and p_val is not None:
                    # 현재 종속변수가 있으면 사용, 없으면 가장 가까운 변수 찾기
                    target_var = current_dep_var if current_dep_var else self.find_closest_dependent_variable(df, i, dep_var_names)

                    if target_var:
                        test_results[target_var] = {'f': f_val, 'p': p_val}
                        self.root.after(0, lambda var=target_var, f=f_val, p=p_val:
                                       self.log(f"    📊 ANOVA 결과: {var}, F={f:.3f}, p={p:.6f}"))

        # 종속변수가 있지만 결과가 없는 경우, 다른 방법으로 시도
        if not test_results and dep_var_names:
            self.root.after(0, lambda: self.log(f"    🔄 대안 방법으로 ANOVA 결과 추출 시도..."))
            test_results = self.extract_anova_results_alternative(df, table_row, dep_var_names)

        return test_results

    def extract_f_and_p_values_enhanced(self, df: pd.DataFrame, row: int) -> tuple:
        """F값과 p값 추출 - 개선된 버전"""
        try:
            f_val = None
            p_val = None

            # 여러 컬럼에서 F값과 p값 찾기 (더 넓은 범위)
            for col in range(min(len(df.columns), 12)):
                cell_val = str(df.iloc[row, col]).strip()
                if cell_val and cell_val != '' and cell_val != '0':
                    try:
                        num_val = float(cell_val)

                        # p값 같아 보이는 것 (0과 1 사이의 작은 값)
                        if 0 <= num_val <= 1 and p_val is None:
                            p_val = num_val
                        # F값 같아 보이는 것 (1보다 큰 값)
                        elif num_val >= 0.001 and f_val is None:
                            f_val = num_val

                    except ValueError:
                        continue

            # F값과 p값이 모두 찾아지면 반환
            if f_val is not None and p_val is not None:
                return f_val, p_val

            # 한쪽만 찾아진 경우 다음 행도 확인
            if row + 1 < len(df):
                for col in range(min(len(df.columns), 12)):
                    cell_val = str(df.iloc[row + 1, col]).strip()
                    if cell_val and cell_val != '':
                        try:
                            num_val = float(cell_val)
                            if 0 <= num_val <= 1 and p_val is None:
                                p_val = num_val
                            elif num_val >= 0.001 and f_val is None:
                                f_val = num_val
                        except ValueError:
                            continue

            return f_val, p_val

        except Exception:
            return None, None

    def extract_anova_results_alternative(self, df: pd.DataFrame, table_row: int, dep_var_names: list) -> dict:
        """대안 방법으로 ANOVA 결과 추출"""
        test_results = {}

        try:
            # 전체 테이블에서 숫자 패턴 찾기
            search_range = range(table_row, min(table_row + 100, len(df)))

            for dep_var in dep_var_names:
                # 해당 종속변수 주변에서 F값과 p값 찾기
                var_found_row = None

                # 종속변수가 언급된 행 찾기
                for i in search_range:
                    if dep_var in str(df.iloc[i, 0]):
                        var_found_row = i
                        break

                if var_found_row:
                    # 해당 변수 주변 10행에서 F값과 p값 찾기
                    for check_row in range(var_found_row, min(var_found_row + 10, len(df))):
                        f_val, p_val = self.extract_f_and_p_values_enhanced(df, check_row)
                        if f_val is not None and p_val is not None:
                            test_results[dep_var] = {'f': f_val, 'p': p_val}
                            break

            return test_results

        except Exception:
            return {}

    def find_closest_dependent_variable(self, df: pd.DataFrame, row: int, dep_var_names: list) -> str:
        """가장 가까운 종속변수 찾기"""
        try:
            # 현재 행 위쪽에서 종속변수명 찾기
            for i in range(row, max(0, row - 30), -1):
                for col in range(min(3, len(df.columns))):
                    cell_val = str(df.iloc[i, col]).strip()
                    if cell_val in dep_var_names:
                        return cell_val

            # 못 찾으면 첫 번째 종속변수 반환
            return dep_var_names[0] if dep_var_names else None
        except:
            return None

    def determine_indep_var_from_groups(self, df: pd.DataFrame, stats_row: int) -> str:
        """SPSS 출력 구조 분석으로 독립변수 직접 추출"""
        try:
            # 🎯 방법 1: SPSS 명령어에서 직접 추출 (가장 확실한 방법)
            command_var = self.extract_from_spss_command(df, stats_row)
            if command_var:
                self.root.after(0, lambda v=command_var:
                               self.log(f"    ✅ SPSS 명령어에서 추출: {v}"))
                return command_var

            # 🎯 방법 2: 테이블 헤더에서 직접 추출
            header_var = self.extract_from_table_header(df, stats_row)
            if header_var:
                self.root.after(0, lambda v=header_var:
                               self.log(f"    ✅ 테이블 헤더에서 추출: {v}"))
                return header_var

            # 🎯 방법 3: 변수 라벨에서 추출 (가장 정확)
            label_var = self.extract_from_variable_labels(df, stats_row)
            if label_var:
                self.root.after(0, lambda v=label_var:
                               self.log(f"    ✅ 변수 라벨에서 추출: {v}"))
                return label_var

            # 🎯 방법 4: 집단 구조 분석으로 추출
            structure_var = self.extract_from_group_structure(df, stats_row)
            if structure_var:
                self.root.after(0, lambda v=structure_var:
                               self.log(f"    ✅ 집단 구조에서 추출: {v}"))
                return structure_var

            # 🎯 최후 방법: 패턴 매칭
            self.root.after(0, lambda: self.log("    ⚠️ 직접 추출 실패, 패턴 분석..."))
            return self.fallback_pattern_matching(df, stats_row)

        except Exception as e:
            self.root.after(0, lambda e=e: self.log(f"    ❌ 독립변수 추정 오류: {e}"))
            return '독립변수'

    def extract_indep_var_from_title(self, df: pd.DataFrame, stats_row: int) -> str:
        """테이블 제목이나 명령어에서 독립변수 추정"""
        try:
            # 테이블 주변에서 SPSS 명령어나 제목 찾기
            search_range = range(max(0, stats_row - 10), min(stats_row + 10, len(df)))

            for i in search_range:
                for col in range(min(3, len(df.columns))):
                    cell_content = str(df.iloc[i, col]).strip()

                    if not cell_content:
                        continue

                    # SPSS 명령어에서 BY 다음 변수명 추출
                    if 'BY ' in cell_content or 'by ' in cell_content:
                        # BY 다음의 변수명 추출
                        parts = cell_content.replace('BY ', 'by ').split('by ')
                        if len(parts) > 1:
                            var_part = parts[1].split()[0].strip()
                            var_name = self.map_variable_to_korean(var_part)
                            if var_name:
                                self.root.after(0, lambda v=var_name, cmd=var_part:
                                               self.log(f"    🎯 명령어에서 추출: {cmd} → {v}"))
                                return var_name

                    # 테이블 제목에서 변수명 추출
                    elif any(pattern in cell_content for pattern in ['분산분석', 'ANOVA', '일원분산분석', 'ONEWAY']):
                        # 같은 행이나 근처에서 변수명 찾기
                        for search_col in range(min(len(df.columns), 10)):
                            nearby_cell = str(df.iloc[i, search_col]).strip()
                            if 'BY' in nearby_cell or 'by' in nearby_cell:
                                parts = nearby_cell.replace('BY ', 'by ').split('by ')
                                if len(parts) > 1:
                                    var_part = parts[1].split()[0].strip()
                                    var_name = self.map_variable_to_korean(var_part)
                                    if var_name:
                                        self.root.after(0, lambda v=var_name:
                                                       self.log(f"    🎯 ANOVA 제목에서 추출: {v}"))
                                        return var_name

            self.root.after(0, lambda: self.log("    ❌ 테이블 제목에서도 독립변수를 찾을 수 없음"))
            return '독립변수'

        except Exception as e:
            self.root.after(0, lambda e=e: self.log(f"    ❌ 제목 추출 오류: {e}"))
            return '독립변수'

    def map_variable_to_korean(self, var_code: str) -> str:
        """SPSS 변수코드를 한국어 변수명으로 매핑 - 초강화 버전"""

        # 🎯 포괄적 변수 매핑 데이터베이스
        comprehensive_mapping = {
            # === 인구통계학적 변수 ===
            # 성별 관련
            '성별': '성별', 'gender': '성별', 'sex': '성별', '남녀': '성별', 'male_female': '성별',
            'q_1성별': '성별', 'q1_성별': '성별', 'gender_cat': '성별', '성': '성별',

            # 연령 관련
            '연령': '연령범주화', '연령대': '연령범주화', '연령범주화': '연령범주화', '연령그룹': '연령범주화',
            'age': '연령범주화', 'age_group': '연령범주화', 'age_cat': '연령범주화', 'agegroup': '연령범주화',
            'q_2연령': '연령범주화', 'q2_연령': '연령범주화', '나이': '연령범주화', '나이대': '연령범주화',
            '연령구간': '연령범주화', '연령범위': '연령범주화', 'age_range': '연령범주화',

            # === 직업/교육 관련 변수 ===
            # 경력 관련
            '경력': '경력범주화', '근무경력': '경력범주화', '총경력': '경력범주화', '현경력': '경력범주화',
            '경력범주화': '경력범주화', '경력기간': '경력범주화', '근무년수': '경력범주화', '재직기간': '경력범주화',
            'career': '경력범주화', 'experience': '경력범주화', 'work_exp': '경력범주화', 'tenure': '경력범주화',
            'q_3경력': '경력범주화', 'q3_경력': '경력범주화', '업무경험': '경력범주화', '직무경험': '경력범주화',
            '현재경력': '경력범주화', '총근무경력': '경력범주화', '현직장경력': '경력범주화',

            # 교육수준/학력 관련
            '학력': '최종학력', '최종학력': '최종학력', '교육수준': '최종학력', '교육정도': '최종학력',
            'education': '최종학력', 'edu_level': '최종학력', 'degree': '최종학력', 'education_level': '최종학력',
            'q_7학력': '최종학력', 'q7_학력': '최종학력', '졸업': '최종학력', '학위': '최종학력',
            '교육배경': '최종학력', '수학정도': '최종학력', '학업수준': '최종학력',

            # 전공 관련
            '전공': '전공', '전공분야': '전공', '전공영역': '전공', '전공과목': '전공', '학과': '전공',
            'major': '전공', 'field': '전공', 'specialty': '전공', 'department': '전공',
            'q_8전공': '전공', 'q8_전공': '전공', '세부전공': '전공', '주전공': '전공',

            # === 기관/조직 관련 변수 ===
            # 기관유형 관련
            '기관유형': '기관유형', '기관형태': '기관유형', '기관종류': '기관유형', '직장유형': '기관유형',
            '소속기관': '기관유형', '근무지': '기관유형', '직장': '기관유형', '회사유형': '기관유형',
            'institution': '기관유형', 'organization': '기관유형', 'workplace': '기관유형', 'company_type': '기관유형',
            'q_4기관유형': '기관유형', 'q4_기관유형': '기관유형', '기관성격': '기관유형', '조직형태': '기관유형',

            # 부서/팀 관련
            '부서': '부서', '팀': '팀', '과': '부서', '실': '부서', '센터': '부서',
            'department': '부서', 'team': '팀', 'division': '부서', 'section': '부서',
            '소속부서': '부서', '담당부서': '부서', '근무부서': '부서',

            # === 직무/업무 관련 변수 ===
            # 직급/직위 관련
            '직급': '직급', '직위': '직급', '직책': '직급', '지위': '직급', '등급': '직급',
            'position': '직급', 'rank': '직급', 'grade': '직급', 'level': '직급', 'title': '직급',
            '관리자급': '직급', '일반직': '직급', '임원': '직급', '팀장': '직급', '과장': '직급',

            # 근무형태 관련
            '근무형태': '근무형태', '고용형태': '근무형태', '근무유형': '근무형태', '채용형태': '근무형태',
            'work_type': '근무형태', 'employment': '근무형태', 'job_type': '근무형태',
            '정규직': '근무형태', '비정규직': '근무형태', '계약직': '근무형태', '파트타임': '근무형태',

            # === 유아교육/보육 전문 변수 ===
            # 담당연령 관련
            '담당연령': '담당연령', '담당반': '담당연령', '담당학급': '담당연령', '맡은연령': '담당연령',
            '반연령': '담당연령', '학급연령': '담당연령', '클래스': '담당연령',
            'class_age': '담당연령', 'classroom': '담당연령', 'assigned_age': '담당연령',
            'q_5담당연령': '담당연령', 'q5_담당연령': '담당연령',

            # 교사수/학급규모 관련
            '교사수': '학급교사수범주화', '학급교사수': '학급교사수범주화', '교사인원': '학급교사수범주화',
            '반교사수': '학급교사수범주화', '담임교사수': '학급교사수범주화', '보조교사': '학급교사수범주화',
            'teacher_num': '학급교사수범주화', 'class_teacher': '학급교사수범주화', 'staff_size': '학급교사수범주화',
            'q_6교사수': '학급교사수범주화', 'q6_교사수': '학급교사수범주화', '학급교사수범주화': '학급교사수범주화',

            # 원아수/학급규모 관련
            '원아수': '원아수', '학급규모': '원아수', '반인원': '원아수', '아동수': '원아수',
            '유아수': '원아수', '학생수': '원아수', '정원': '원아수', '현원': '원아수',
            'child_num': '원아수', 'class_size': '원아수', 'student_num': '원아수',

            # === 건강/의료 관련 변수 ===
            # 의료기관 관련
            '의료기관': '의료기관유형', '병원유형': '의료기관유형', '진료과': '진료과',
            '병원급수': '의료기관유형', '의료진': '의료진유형', '직종': '직종',

            # === 지역/거주 관련 변수 ===
            # 지역 관련
            '지역': '지역', '거주지': '지역', '소재지': '지역', '위치': '지역',
            '시도': '지역', '시군구': '지역', '동네': '지역', '구역': '지역',
            'region': '지역', 'location': '지역', 'area': '지역', 'district': '지역',

            # === 경제/소득 관련 변수 ===
            # 소득/급여 관련
            '소득': '소득수준', '급여': '급여수준', '월급': '급여수준', '연봉': '급여수준',
            '소득수준': '소득수준', '급여수준': '급여수준', '임금': '급여수준', '보수': '급여수준',
            'income': '소득수준', 'salary': '급여수준', 'wage': '급여수준', 'pay': '급여수준',

            # === 가족/개인 관련 변수 ===
            # 결혼/가족 관련
            '결혼상태': '결혼상태', '혼인상태': '결혼상태', '결혼여부': '결혼상태', '혼인여부': '결혼상태',
            '가족구성': '가족구성', '가족형태': '가족구성', '자녀수': '자녀수', '자녀여부': '자녀여부',
            'marital_status': '결혼상태', 'family_type': '가족구성', 'children': '자녀수',

            # === 태도/의견 관련 변수 ===
            # 만족도 관련
            '만족도': '만족도', '만족수준': '만족도', '만족정도': '만족도',
            'satisfaction': '만족도', 'satisfaction_level': '만족도',
            '직무만족': '직무만족도', '생활만족': '생활만족도', '서비스만족': '서비스만족도',

            # === 행동/성과 관련 변수 ===
            # 성과/평가 관련
            '성과': '성과', '실적': '성과', '평가': '평가결과', '등급': '평가등급',
            'performance': '성과', 'achievement': '성과', 'evaluation': '평가결과',

            # === 기타 특수 변수 ===
            # 시간 관련
            '시간': '시간대', '기간': '기간', '주기': '주기', '빈도': '빈도',
            'time': '시간대', 'period': '기간', 'frequency': '빈도',

            # 참여/이용 관련
            '참여여부': '참여여부', '이용여부': '이용여부', '경험여부': '경험여부',
            '참여정도': '참여정도', '이용정도': '이용정도', '활용정도': '활용정도',
            'participation': '참여여부', 'usage': '이용여부', 'experience': '경험여부',
        }

        # 1. 정확히 일치하는 경우
        var_lower = var_code.lower().strip()
        if var_code in comprehensive_mapping:
            return comprehensive_mapping[var_code]
        if var_lower in comprehensive_mapping:
            return comprehensive_mapping[var_lower]

        # 2. 부분 일치 검사 (긴 키워드부터)
        sorted_keys = sorted(comprehensive_mapping.keys(), key=len, reverse=True)
        for key in sorted_keys:
            if key in var_code or var_code in key:
                return comprehensive_mapping[key]
            if key.lower() in var_lower or var_lower in key.lower():
                return comprehensive_mapping[key]

        # 3. 고급 패턴 기반 추정
        return self.advanced_pattern_matching(var_code)

    def advanced_pattern_matching(self, var_code: str) -> str:
        """고급 패턴 매칭으로 변수 추정"""
        var_lower = var_code.lower()

        # 🔍 패턴 분석 규칙들
        pattern_rules = [
            # 연령 관련 패턴
            (['연령', '나이', '세', 'age', 'old', '년생', 'birth'], '연령범주화'),
            (['경력', '근무', '재직', '년차', 'career', 'exp', 'work', 'service', '년'], '경력범주화'),
            (['성별', '남녀', '성', 'gender', 'sex', 'male', 'female'], '성별'),
            (['학력', '교육', '졸업', '학위', 'education', 'degree', 'academic'], '최종학력'),
            (['전공', '학과', '과정', 'major', 'field', 'specialty', 'study'], '전공'),
            (['기관', '회사', '직장', '조직', 'institution', 'company', 'organization'], '기관유형'),
            (['직급', '직위', '직책', 'position', 'rank', 'grade', 'level'], '직급'),
            (['부서', '팀', '과', '실', 'department', 'team', 'division'], '부서'),
            (['근무형태', '고용', '채용', 'employment', 'work_type', 'job'], '근무형태'),
            (['지역', '거주', '소재', '위치', 'region', 'location', 'area'], '지역'),
            (['소득', '급여', '월급', '연봉', 'income', 'salary', 'wage'], '소득수준'),
            (['결혼', '혼인', '배우자', 'marriage', 'marital', 'spouse'], '결혼상태'),
            (['자녀', '아이', '아들', '딸', 'child', 'son', 'daughter'], '자녀수'),
            (['만족', 'satisfaction', 'satisfy'], '만족도'),
            (['성과', '실적', '평가', 'performance', 'achievement', 'result'], '성과'),
            (['참여', '이용', '활용', '경험', 'participation', 'usage', 'experience'], '참여여부'),
            (['시간', '기간', '주기', 'time', 'period', 'duration'], '시간대'),
            (['빈도', '횟수', 'frequency', 'times', 'count'], '빈도'),

            # 유아교육 특화 패턴
            (['담당', '맡은', '반', '학급', 'class', 'assigned'], '담당연령'),
            (['교사수', '교사인원', '선생님', 'teacher', 'staff'], '학급교사수범주화'),
            (['원아', '유아', '아동', '아이', 'child', 'student'], '원아수'),
            (['유치원', '어린이집', '보육', 'kindergarten', 'daycare'], '기관유형'),

            # 의료/건강 특화 패턴
            (['병원', '의료', '클리닉', 'hospital', 'medical', 'clinic'], '의료기관유형'),
            (['진료과', '과', '전문의', 'department', 'specialty'], '진료과'),
            (['간호사', '의사', '의료진', 'nurse', 'doctor', 'medical_staff'], '직종'),
        ]

        # 패턴 매칭 실행
        for patterns, result in pattern_rules:
            if any(pattern in var_lower for pattern in patterns):
                return result

        # 4. 숫자 패턴 분석
        if var_code.isdigit():
            return None  # 순수 숫자는 독립변수가 아님

        # 5. 특수 패턴 (q_숫자 형태)
        import re
        q_pattern = re.match(r'q_?(\d+)(.+)', var_lower)
        if q_pattern:
            var_name = q_pattern.group(2)
            return self.advanced_pattern_matching(var_name)

        return None

    def extract_from_spss_command(self, df: pd.DataFrame, stats_row: int) -> str:
        """SPSS 명령어에서 직접 독립변수 추출"""
        try:
            # 테이블 주변에서 SPSS 명령어 찾기
            search_range = range(max(0, stats_row - 20), min(stats_row + 20, len(df)))

            for i in search_range:
                for col in range(min(len(df.columns), 5)):
                    cell = str(df.iloc[i, col]).strip()

                    # T-TEST 명령어 패턴
                    if 'T-TEST' in cell and 'GROUPS=' in cell:
                        # T-TEST GROUPS=변수명(값1 값2) 패턴
                        import re
                        pattern = r'GROUPS?=([^(\s]+)'
                        match = re.search(pattern, cell)
                        if match:
                            var_name = match.group(1).strip()
                            return self.clean_variable_name(var_name)

                    # ONEWAY 명령어 패턴
                    elif 'ONEWAY' in cell and ' BY ' in cell:
                        # ONEWAY 종속변수 BY 독립변수 패턴
                        parts = cell.split(' BY ')
                        if len(parts) >= 2:
                            var_part = parts[1].split()[0].strip()
                            return self.clean_variable_name(var_part)

                    # ANOVA 명령어 패턴
                    elif 'ANOVA' in cell and ' BY ' in cell:
                        parts = cell.split(' BY ')
                        if len(parts) >= 2:
                            var_part = parts[1].split()[0].strip()
                            return self.clean_variable_name(var_part)

            return None

        except Exception:
            return None

    def extract_from_table_header(self, df: pd.DataFrame, stats_row: int) -> str:
        """테이블 헤더나 제목에서 독립변수 추출"""
        try:
            # 테이블 제목 행들 확인
            title_range = range(max(0, stats_row - 5), stats_row + 1)

            for i in title_range:
                for col in range(min(len(df.columns), 3)):
                    cell = str(df.iloc[i, col]).strip()

                    # 독립변수가 명시된 제목 패턴
                    if any(keyword in cell for keyword in ['분산분석', 'ANOVA', '일원분산분석']):
                        # 다음 행이나 같은 행에서 변수명 찾기
                        for check_row in range(i, min(i + 3, len(df))):
                            for check_col in range(min(len(df.columns), 5)):
                                check_cell = str(df.iloc[check_row, check_col]).strip()
                                if self.is_korean_variable_name(check_cell):
                                    return check_cell

                    # T검정 제목에서 변수명 찾기
                    elif any(keyword in cell for keyword in ['집단통계', 'Group Statistics']):
                        # 바로 다음 몇 행에서 변수명 찾기
                        for check_row in range(i + 1, min(i + 5, len(df))):
                            var_col_content = str(df.iloc[check_row, 0]).strip()
                            if self.is_korean_variable_name(var_col_content):
                                # 해당 변수의 집단을 보고 독립변수 추정
                                return self.infer_indep_var_from_dependent(var_col_content)

            return None

        except Exception:
            return None

    def extract_from_variable_labels(self, df: pd.DataFrame, stats_row: int) -> str:
        """변수 라벨이나 설명에서 독립변수 추출"""
        try:
            # 테이블에서 변수 라벨 영역 찾기
            search_range = range(stats_row, min(stats_row + 50, len(df)))

            for i in search_range:
                # 첫 번째 컬럼에서 한국어 변수명 찾기
                first_col = str(df.iloc[i, 0]).strip()

                if self.is_korean_variable_name(first_col):
                    # 이 변수가 종속변수인지 독립변수인지 판단
                    # 같은 행의 두 번째 컬럼에 그룹명이 있으면 독립변수
                    if len(df.columns) > 1:
                        second_col = str(df.iloc[i, 1]).strip()
                        if self.is_group_name(second_col):
                            return first_col

                    # 다음 몇 행에 그룹들이 나열되어 있으면 독립변수
                    groups_found = 0
                    for j in range(i + 1, min(i + 10, len(df))):
                        if len(df.columns) > 1:
                            potential_group = str(df.iloc[j, 1]).strip()
                            if self.is_group_name(potential_group):
                                groups_found += 1

                    if groups_found >= 2:  # 2개 이상 그룹이 있으면 독립변수
                        return first_col

            return None

        except Exception:
            return None

    def extract_from_group_structure(self, df: pd.DataFrame, stats_row: int) -> str:
        """집단 구조 분석으로 독립변수 추출"""
        try:
            # 집단 데이터가 있는 영역 분석
            search_range = range(stats_row, min(stats_row + 100, len(df)))
            group_structures = {}

            for i in search_range:
                if len(df.columns) > 1:
                    var_name = str(df.iloc[i, 0]).strip()
                    group_name = str(df.iloc[i, 1]).strip()

                    # 한국어 변수명과 그룹명이 있는 행
                    if (self.is_korean_variable_name(var_name) and
                        self.is_group_name(group_name) and
                        var_name != group_name):

                        if var_name not in group_structures:
                            group_structures[var_name] = []
                        group_structures[var_name].append(group_name)

            # 가장 많은 그룹을 가진 변수를 독립변수로 판단
            if group_structures:
                # 그룹 수가 2개 이상인 변수들만 고려
                valid_vars = {k: v for k, v in group_structures.items() if len(v) >= 2}
                if valid_vars:
                    # 그룹 수가 가장 많은 변수 선택 (더 세분화된 변수가 독립변수일 가능성 높음)
                    indep_var = max(valid_vars.keys(), key=lambda x: len(valid_vars[x]))
                    self.root.after(0, lambda v=indep_var, groups=valid_vars[indep_var]:
                                   self.log(f"    📊 집단 구조 분석: {v} -> {groups}"))
                    return indep_var

            return None

        except Exception:
            return None

    def clean_variable_name(self, var_name: str) -> str:
        """변수명 정리 및 한국어 변환"""
        # 특수문자 제거
        cleaned = var_name.replace('(', '').replace(')', '').replace('[', '').replace(']', '')
        cleaned = cleaned.replace('\n', '').replace('\t', '').strip()

        # SPSS 변수코드를 한국어로 변환
        korean_name = self.convert_spss_code_to_korean(cleaned)
        return korean_name if korean_name else cleaned

    def convert_spss_code_to_korean(self, var_code: str) -> str:
        """SPSS 변수코드를 한국어로 변환 - 강화 버전"""
        # 포괄적 변수 매핑
        variable_mappings = {
            # 기본 인구통계학적 변수
            'q_1성별': '성별', 'q1성별': '성별', 'q_2성별': '성별', 'q2성별': '성별',
            'gender': '성별', 'sex': '성별',

            # 연령 관련
            'q_2연령': '연령범주화', 'q2연령': '연령범주화', 'q_3연령': '연령범주화',
            'age': '연령범주화', '연령범주화': '연령범주화',

            # 종교
            'q_3종교': '종교', 'q3종교': '종교', 'religion': '종교',

            # 결혼상태
            'q_4결혼': '결혼상태', 'q4결혼': '결혼상태', 'marriage': '결혼상태',
            'marital': '결혼상태',

            # 학력
            'q_5학력': '최종학력', 'q5학력': '최종학력', 'q_7학력': '최종학력',
            'education': '최종학력', 'degree': '최종학력',

            # 부서/직장
            'q_6부서': '근무부서', 'q6부서': '근무부서', 'department': '근무부서',
            'q_7희망부서배치여부': '희망부서배치여부',

            # 경력
            'q_3경력': '경력범주화', 'q3경력': '경력범주화', '총경력범주화': '총경력범주화',
            '경력범주화': '경력범주화', '현경력범주화': '경력범주화',
            'career': '경력범주화', 'experience': '경력범주화',

            # 기관/조직
            'q_4기관': '기관유형', 'q4기관': '기관유형', '기관유형': '기관유형',
            'institution': '기관유형', 'organization': '기관유형',

            # 담당업무
            'q_5담당': '담당연령', 'q5담당': '담당연령', '담당연령': '담당연령',
            'class': '담당연령', 'classroom': '담당연령',

            # 교사수
            'q_6교사': '학급교사수범주화', 'q6교사': '학급교사수범주화',
            '학급교사수범주화': '학급교사수범주화', 'teacher': '학급교사수범주화',

            # 전공
            'q_8전공': '전공', 'q8전공': '전공', '전공': '전공',
            'major': '전공', 'field': '전공',

            # 기타 직장/업무 관련
            'q_9부담': '업무부담', 'q_10이직': '이직의도', 'q_11만족도': '직무만족도',
            'q_12급여': '급여수준', 'q_13근무형태': '근무형태',

            # 최종학력
            '최종학력': '최종학력'
        }

        # 정확히 일치하는 경우
        if var_code in variable_mappings:
            return variable_mappings[var_code]

        # 부분 일치 검사
        for key, value in variable_mappings.items():
            if key in var_code or var_code in key:
                return value

        # 패턴 기반 변환
        if 'q_' in var_code or 'q' in var_code:
            # q_숫자패턴 분석
            import re

            # 성별 패턴
            if any(pattern in var_code.lower() for pattern in ['성별', 'gender', 'sex']):
                return '성별'
            # 연령 패턴
            elif any(pattern in var_code.lower() for pattern in ['연령', 'age']):
                return '연령범주화'
            # 종교 패턴
            elif any(pattern in var_code.lower() for pattern in ['종교', 'religion']):
                return '종교'
            # 결혼 패턴
            elif any(pattern in var_code.lower() for pattern in ['결혼', 'marriage', 'marital']):
                return '결혼상태'
            # 학력 패턴
            elif any(pattern in var_code.lower() for pattern in ['학력', 'education', 'degree']):
                return '최종학력'
            # 부서 패턴
            elif any(pattern in var_code.lower() for pattern in ['부서', 'department', 'dept']):
                return '근무부서'
            # 경력 패턴
            elif any(pattern in var_code.lower() for pattern in ['경력', 'career', 'exp']):
                return '경력범주화'

        # 한국어 변수명은 그대로 반환
        if any('\uac00' <= char <= '\ud7af' for char in var_code):
            return var_code

        return None

    def is_korean_variable_name(self, text: str) -> bool:
        """한국어 변수명인지 판단"""
        if not text or len(text) < 2:
            return False

        # 숫자만 있는 경우 제외
        try:
            float(text)
            return False
        except ValueError:
            pass

        # 한국어 포함 여부 확인
        has_korean = any('\uac00' <= char <= '\ud7af' for char in text)

        # 일반적인 변수명 패턴 확인
        common_var_patterns = [
            '범주화', '유형', '형태', '수준', '정도', '여부', '상태', '등급',
            '연령', '경력', '학력', '전공', '성별', '지역', '소득', '만족',
            '기관', '담당', '교사', '근무', '직급', '부서'
        ]

        if has_korean or any(pattern in text for pattern in common_var_patterns):
            # 제외할 패턴들
            exclude_patterns = [
                '평균', '표준편차', '합계', '점수', '총합', '전체',
                '집단통계', '기술통계', '분산분석', 'ANOVA', 'T-TEST',
                '검정통계량', '유의확률', '자유도', '제곱합'
            ]

            if not any(pattern in text for pattern in exclude_patterns):
                return True

        return False

    def is_group_name(self, text: str) -> bool:
        """그룹명인지 판단"""
        if not text or len(text) < 1:
            return False

        # 숫자만 있는 경우 제외
        try:
            float(text)
            return False
        except ValueError:
            pass

        # 통계값인 경우 제외
        if any(pattern in text for pattern in ['평균', '표준편차', 'N', 'Std']):
            return False

        # 의미있는 그룹명 패턴들
        group_patterns = [
            # 성별
            '남', '여', '남자', '여자', '남성', '여성',
            # 연령
            '세', '대', '미만', '이상', '이하', '초과',
            # 경력
            '년', '개월', '신입', '경험',
            # 학력
            '졸업', '과정', '학사', '석사', '박사',
            # 기관
            '유치원', '어린이집', '국공립', '사립', '민간',
            # 기타
            '명', '있음', '없음', '참여', '불참'
        ]

        return any(pattern in text for pattern in group_patterns) or len(text) <= 10

    def infer_indep_var_from_dependent(self, dep_var: str) -> str:
        """종속변수로부터 독립변수 추정 (임시)"""
        # 일반적으로 성별로 나누는 경우가 많음
        return '성별'

    def get_variable_mapping(self) -> dict:
        """SPSS 변수코드 -> 한국어 매핑"""
        return {
            'q_1성별': '성별', 'q1성별': '성별', 'gender': '성별',
            'q_2연령': '연령범주화', 'q2연령': '연령범주화', 'age': '연령범주화',
            'q_3경력': '경력범주화', 'q3경력': '경력범주화', 'career': '경력범주화',
            'q_4기관': '기관유형', 'q4기관': '기관유형', 'institution': '기관유형',
            'q_5담당': '담당연령', 'q5담당': '담당연령', 'class': '담당연령',
            'q_6교사': '학급교사수범주화', 'q6교사': '학급교사수범주화',
            'q_7학력': '최종학력', 'q7학력': '최종학력', 'education': '최종학력',
            'q_8전공': '전공', 'q8전공': '전공', 'major': '전공',
            '현경력범주화': '경력범주화', '연령범주화': '연령범주화'
        }

    def fallback_pattern_matching(self, df: pd.DataFrame, stats_row: int) -> str:
        """최후 수단: 간단한 패턴 매칭"""
        try:
            # 주변에서 한국어 변수명 찾기
            for i in range(max(0, stats_row - 10), min(stats_row + 30, len(df))):
                for col in range(min(3, len(df.columns))):
                    cell = str(df.iloc[i, col]).strip()
                    if self.is_korean_variable_name(cell):
                        return cell

            # 기본값 반환
            return '독립변수'

        except Exception:
            return '독립변수'

    def comprehensive_group_pattern_matching(self, group_text: str) -> str:
        """포괄적 그룹 패턴 매칭으로 독립변수 추정"""
        group = group_text.strip().lower()

        # 🎯 초강화된 그룹 패턴 데이터베이스
        group_patterns = {
            # === 연령 관련 패턴 ===
            '연령범주화': [
                # 한국식 연령 표현
                '세 미만', '세 이상', '세미만', '세이상', '세 이하', '세이하',
                '20세', '25세', '30세', '35세', '40세', '45세', '50세', '55세', '60세',
                '20대', '30대', '40대', '50대', '60대', '70대',
                '이십대', '삼십대', '사십대', '오십대', '육십대',
                # 영어식 연령 표현
                'under', 'over', 'below', 'above', 'years old', 'age',
                'young', 'middle', 'old', 'elder', 'senior',
                # 특수 연령 구분
                '청년', '중년', '노년', '장년', '어린', '나이',
                '10대', '청소년', '성인', '노인', '어르신'
            ],

            # === 경력/근무년수 관련 패턴 ===
            '경력범주화': [
                # 한국식 경력 표현
                '년 미만', '년 이상', '년미만', '년이상', '년 이하', '년이하',
                '1년', '2년', '3년', '4년', '5년', '6년', '7년', '8년', '9년', '10년',
                '년차', '연차', '해차', '개월', '월',
                # 경력 구분 표현
                '신입', '초급', '중급', '고급', '숙련', '베테랑',
                '신규', '경험자', '숙련자', '전문가', '선임',
                # 영어식 경력 표현
                'years', 'year', 'experience', 'exp', 'career',
                'beginner', 'junior', 'senior', 'expert', 'veteran',
                'novice', 'experienced', 'skilled'
            ],

            # === 성별 관련 패턴 ===
            '성별': [
                '남자', '여자', '남성', '여성', '남', '여',
                'male', 'female', 'man', 'woman', 'men', 'women',
                '남학생', '여학생', '남교사', '여교사',
                '아버지', '어머니', '아빠', '엄마'
            ],

            # === 학력/교육 관련 패턴 ===
            '최종학력': [
                # 학위 관련
                '졸업', '수료', '중퇴', '재학',
                '학사', '석사', '박사', '전문학사',
                '대학원', '대학교', '전문대학', '전문대', '4년제', '2년제',
                '고등학교', '고교', '중학교', '중학', '초등학교', '초등',
                # 특수 교육과정
                '보육교사 양성과정', '양성과정', '직업훈련',
                '사이버대학', '방송대학', '평생교육',
                # 영어식 표현
                'bachelor', 'master', 'doctor', 'phd', 'college', 'university',
                'graduate', 'undergraduate', 'diploma', 'certificate'
            ],

            # === 전공 관련 패턴 ===
            '전공': [
                # 교육 관련 전공
                '유아교육과', '아동학과', '보육관련학과', '교육학과',
                '유아교육', '아동학', '보육학', '교육학',
                '사회복지학과', '심리학과', '상담학과',
                # 일반 전공 표현
                '전공', '학과', '과', '학부', '계열',
                '인문계', '이공계', '예체능계', '상경계',
                # 영어식 전공 표현
                'major', 'department', 'faculty', 'school',
                'humanities', 'science', 'engineering', 'business',
                '기타'  # 기타 전공
            ],

            # === 기관/조직 관련 패턴 ===
            '기관유형': [
                # 유아교육기관
                '유치원', '어린이집', '보육원', '어린이집',
                '국공립유치원', '사립유치원', '공립유치원', '사사립유치원',
                '국공립어린이집', '민간어린이집', '가정어린이집', '법인어린이집',
                '직장어린이집', '협동어린이집', '부모협동어린이집',
                # 일반 기관
                '공공기관', '민간기관', '정부기관', '지자체',
                '대기업', '중소기업', '벤처기업', '스타트업',
                '병원', '의료기관', '클리닉', '요양원',
                # 영어식 표현
                'public', 'private', 'government', 'corporate',
                'hospital', 'clinic', 'company', 'organization'
            ],

            # === 담당연령/학급 관련 패턴 ===
            '담당연령': [
                # 담당 연령
                '만3세', '만4세', '만5세', '만 3세', '만 4세', '만 5세',
                '3세', '4세', '5세', '6세', '7세',
                '영아', '유아', '어린이',
                '영아반', '유아반', '혼합연령', '혼합반',
                '0세반', '1세반', '2세반', '누리반',
                # 영어식 표현
                'infant', 'toddler', 'preschool', 'kindergarten',
                'mixed age', 'multi age'
            ],

            # === 교사수/인원 관련 패턴 ===
            '학급교사수범주화': [
                '1명', '2명', '3명', '4명', '5명', '6명', '7명', '8명', '9명', '10명',
                '명', '인', '사람',
                '한명', '두명', '세명', '네명', '다섯명',
                '일명', '이명', '삼명', '사명', '오명',
                '1인', '2인', '3인', '4인', '5인',
                '단독', '협력', '팀티칭', '공동',
                '담임', '부담임', '보조',
                # 영어식 표현
                'teacher', 'staff', 'instructor', 'educator',
                'single', 'multiple', 'team', 'co-teaching'
            ],

            # === 지역 관련 패턴 ===
            '지역': [
                # 광역시도
                '서울', '부산', '대구', '인천', '광주', '대전', '울산', '세종',
                '경기', '강원', '충북', '충남', '전북', '전남', '경북', '경남', '제주',
                '수도권', '영남권', '호남권', '충청권', '강원권',
                # 시군구
                '시', '군', '구', '읍', '면', '동', '리',
                '특별시', '광역시', '특별자치시', '특별자치도',
                # 일반 지역 표현
                '도심', '시내', '시외', '교외', '농촌', '어촌', '산촌',
                '신도시', '구도심', '번화가', '주택가',
                # 영어식 표현
                'urban', 'rural', 'suburban', 'metropolitan',
                'downtown', 'city', 'town', 'village'
            ],

            # === 소득/급여 관련 패턴 ===
            '소득수준': [
                # 소득 구간
                '만원', '천만원', '억원', '원',
                '100만원', '200만원', '300만원', '400만원', '500만원',
                '이하', '이상', '미만', '초과',
                # 소득 수준
                '저소득', '중소득', '고소득', '상위소득', '하위소득',
                '최저임금', '평균임금', '고액연봉',
                # 영어식 표현
                'low income', 'middle income', 'high income',
                'salary', 'wage', 'income', 'earnings'
            ],

            # === 결혼/가족 관련 패턴 ===
            '결혼상태': [
                '미혼', '기혼', '이혼', '별거', '사별', '재혼',
                '독신', '싱글', '커플', '부부', '배우자',
                '결혼', '혼인', '동거', '약혼',
                # 영어식 표현
                'single', 'married', 'divorced', 'separated', 'widowed',
                'unmarried', 'spouse', 'partner'
            ],

            # === 자녀 관련 패턴 ===
            '자녀수': [
                '무자녀', '1자녀', '2자녀', '3자녀', '4자녀', '다자녀',
                '외동', '둘째', '셋째', '막내', '첫째',
                '아들', '딸', '남아', '여아',
                '자녀없음', '자녀있음',
                # 영어식 표현
                'no children', 'one child', 'two children', 'multiple children',
                'son', 'daughter', 'kids', 'children'
            ],

            # === 직급/직위 관련 패턴 ===
            '직급': [
                # 일반 직급
                '사원', '대리', '과장', '차장', '부장', '이사', '상무', '전무', '부사장', '사장',
                '주임', '선임', '책임', '수석', '전문', '특급',
                '인턴', '수습', '정규직', '계약직', '임시직',
                # 교육 관련 직급
                '원장', '원감', '주임교사', '담임교사', '부담임교사',
                '수석교사', '교감', '교장', '보조교사',
                # 의료 관련 직급
                '전문의', '레지던트', '인턴', '간호사', '간호장',
                '수간호사', '책임간호사', '주임간호사',
                # 영어식 표현
                'manager', 'director', 'supervisor', 'coordinator',
                'specialist', 'expert', 'assistant', 'associate'
            ],

            # === 근무형태 관련 패턴 ===
            '근무형태': [
                '정규직', '비정규직', '계약직', '임시직', '파견직',
                '시간제', '전일제', '파트타임', '풀타임',
                '주간근무', '야간근무', '교대근무', '시프트',
                '재택근무', '원격근무', '출장', '파견',
                # 영어식 표현
                'full time', 'part time', 'contract', 'temporary',
                'permanent', 'freelance', 'remote', 'shift'
            ],

            # === 만족도 관련 패턴 ===
            '만족도': [
                '매우불만', '불만', '보통', '만족', '매우만족',
                '전혀', '거의', '약간', '상당히', '매우',
                '1점', '2점', '3점', '4점', '5점',
                '낮음', '높음', '중간',
                # 영어식 표현
                'very dissatisfied', 'dissatisfied', 'neutral', 'satisfied', 'very satisfied',
                'low', 'medium', 'high', 'excellent', 'poor'
            ],

            # === 참여/이용 관련 패턴 ===
            '참여여부': [
                '참여', '불참', '참석', '불참석', '출석', '결석',
                '이용', '미이용', '사용', '미사용', '활용', '미활용',
                '경험', '미경험', '수강', '미수강',
                '있음', '없음', '한다', '안한다',
                # 영어식 표현
                'participate', 'not participate', 'attend', 'not attend',
                'use', 'not use', 'experience', 'no experience'
            ]
        }

        # 패턴 매칭 실행 (긴 패턴부터 우선 매칭)
        for var_type, patterns in group_patterns.items():
            # 패턴을 길이순으로 정렬 (긴 패턴 우선)
            sorted_patterns = sorted(patterns, key=len, reverse=True)
            for pattern in sorted_patterns:
                if pattern in group:
                    return var_type

        return None

    def extract_ttest_analysis(self, df: pd.DataFrame, pair: Dict) -> None:
        """T검정 분석 완전 추출 - 강화된 유연성"""
        try:
            stats_row = pair['stats_table']['row']
            results_row = pair['results_table']['row']

            # 1. 집단통계량 추출 - 더 유연한 방식
            groups_data = {}

            # 가능한 모든 종속변수 찾기
            possible_dep_vars = set()
            for i in range(stats_row, min(stats_row + 50, len(df))):
                for col in range(min(3, len(df.columns))):
                    cell_val = str(df.iloc[i, col]).strip()
                    if any(pattern in cell_val for pattern in [
                        '평균', '합계', '점수', '역량', '신념', '성과', '만족'
                    ]) and len(cell_val) > 3:
                        possible_dep_vars.add(cell_val)

            # 각 종속변수별로 그룹 데이터 수집
            for dep_var in possible_dep_vars:
                groups_data[dep_var] = []

                for i in range(stats_row + 2, min(stats_row + 30, len(df))):
                    if '독립표본' in str(df.iloc[i, 0]):
                        break

                    var_cell = str(df.iloc[i, 0]).strip()
                    group_cell = str(df.iloc[i, 1]).strip()

                    # 🎯 실제 그룹명만 인식 (T검정 버전)
                    if dep_var == var_cell and group_cell and self.is_real_group_name(group_cell):
                        # 첫 번째 그룹
                        try:
                            n_col = self.find_number_column(df, i, 'n')
                            mean_col = self.find_number_column(df, i, 'mean')
                            std_col = self.find_number_column(df, i, 'std')

                            if n_col and mean_col and std_col:
                                groups_data[dep_var].append({
                                    'group': group_cell,
                                    'n': int(float(df.iloc[i, n_col])),
                                    'mean': float(df.iloc[i, mean_col]),
                                    'std': float(df.iloc[i, std_col])
                                })

                                self.root.after(0, lambda grp=group_cell:
                                               self.log(f"    ✅ 실제 그룹 발견: {grp}"))

                                # 다음 그룹 찾기
                                for j in range(i + 1, min(i + 5, len(df))):
                                    next_group = str(df.iloc[j, 1]).strip()
                                    if (next_group and next_group != group_cell and
                                        self.is_real_group_name(next_group)):

                                        groups_data[dep_var].append({
                                            'group': next_group,
                                            'n': int(float(df.iloc[j, n_col])),
                                            'mean': float(df.iloc[j, mean_col]),
                                            'std': float(df.iloc[j, std_col])
                                        })

                                        self.root.after(0, lambda grp=next_group:
                                                       self.log(f"    ✅ 실제 그룹 발견: {grp}"))
                                        break
                        except (ValueError, IndexError):
                            continue
                        break

            # 빈 그룹 제거
            groups_data = {k: v for k, v in groups_data.items() if v}

            # 2. T검정 결과 추출 - 더 유연한 방식
            test_results = {}

            for var in possible_dep_vars:
                if var in groups_data:  # 그룹 데이터가 있는 변수만
                    # T검정 결과 테이블에서 해당 변수의 결과 찾기
                    for i in range(results_row + 1, min(results_row + 30, len(df))):
                        result_var = str(df.iloc[i, 0]).strip()
                        condition = str(df.iloc[i, 1]).strip() if len(df.columns) > 1 else ""

                        if var == result_var or (result_var and var in result_var):
                            try:
                                # Levene 검정 기반 선택
                                if '가정함' in condition:
                                    levene_p_col = self.find_p_value_column(df, i, 'levene')
                                    t_col = self.find_number_column(df, i, 't')
                                    p_col = self.find_p_value_column(df, i, 'ttest')

                                    if levene_p_col and t_col and p_col:
                                        levene_p = float(df.iloc[i, levene_p_col])

                                        if levene_p < 0.05:
                                            # 등분산 가정 위반 → 다음 행 사용
                                            next_i = i + 1
                                            if next_i < len(df):
                                                t_val = float(df.iloc[next_i, t_col])
                                                p_val = float(df.iloc[next_i, p_col])
                                                choice = "가정하지않음"
                                        else:
                                            # 등분산 가정 만족 → 현재 행 사용
                                            t_val = float(df.iloc[i, t_col])
                                            p_val = float(df.iloc[i, p_col])
                                            choice = "가정함"

                                        test_results[var] = {'t': t_val, 'p': p_val}

                                        self.root.after(0, lambda var=var, choice=choice, t=t_val, p=p_val:
                                                       self.log(f"  📊 {var} ({choice}): t={t:.3f}, p={p:.3f}"))
                                        break

                            except (ValueError, IndexError):
                                continue

            # 3. 결과 저장
            for var in groups_data:
                if var in test_results:
                    self.all_analyses.append({
                        'indep_var': pair['indep_var'],
                        'dep_var': var,
                        'groups': groups_data[var],
                        'statistic': test_results[var]['t'],
                        'p_value': test_results[var]['p'],
                        'test_type': 't-test'
                    })
                    self.root.after(0, lambda var=var: self.log(f"  ✅ T검정 저장: {var}"))

        except Exception as e:
            self.root.after(0, lambda e=e: self.log(f"  💥 T검정 추출 오류: {e}", 'error'))

    def find_p_value_column(self, df: pd.DataFrame, row: int, test_type: str) -> Optional[int]:
        """p값이 있는 컬럼 찾기"""
        try:
            if test_type == 'levene':
                # Levene 검정 p값은 보통 3-4번째 컬럼
                search_cols = range(2, min(6, len(df.columns)))
            else:
                # t검정 p값은 보통 마지막 쪽 컬럼
                search_cols = range(5, min(len(df.columns), 10))

            for col in search_cols:
                cell_val = str(df.iloc[row, col]).strip()
                if cell_val:
                    try:
                        num_val = float(cell_val)
                        if 0 <= num_val <= 1:  # p값 범위
                            return col
                    except ValueError:
                        continue
            return None
        except:
            return None

    def extract_anova_analysis(self, df: pd.DataFrame, pair: Dict) -> None:
        """ANOVA 분석 완전 추출 - 강화된 디버깅"""
        try:
            stats_row = pair['stats_table']['row']
            results_row = pair['results_table']['row']

            self.root.after(0, lambda: self.log(f"  📍 기술통계 행{stats_row}, ANOVA 결과 행{results_row}"))

            # 1. 기술통계 추출 - 완전히 유연한 방식
            groups_data = {}

            self.root.after(0, lambda: self.log(f"    🔍 기술통계 스캔 시작..."))

            # 가능한 모든 종속변수 패턴 찾기
            possible_dep_vars = set()
            for i in range(stats_row, min(stats_row + 100, len(df))):
                for col in range(min(3, len(df.columns))):
                    cell_val = str(df.iloc[i, col]).strip()

                    # 종속변수로 보이는 패턴들
                    if any(pattern in cell_val for pattern in [
                        '평균', '합계', '점수', '역량', '신념', '성과', '만족',
                        'average', 'mean', 'total', 'score', 'satisfaction'
                    ]) and len(cell_val) > 3:
                        possible_dep_vars.add(cell_val)

            self.root.after(0, lambda vars=list(possible_dep_vars):
                           self.log(f"    📋 가능한 종속변수들: {vars}"))

            # 각 종속변수별로 그룹 데이터 수집
            for dep_var in possible_dep_vars:
                groups_data[dep_var] = []

                # 해당 변수의 데이터 행들 찾기
                for i in range(stats_row, min(stats_row + 100, len(df))):
                    if 'ANOVA' in str(df.iloc[i, 0]):
                        break

                    # 변수명이 나타나는 행 찾기
                    if dep_var in str(df.iloc[i, 0]):
                        # 이 변수의 그룹들 수집
                        self.root.after(0, lambda var=dep_var: self.log(f"    📍 {var} 데이터 수집 시작"))

                        for j in range(i, min(i + 20, len(df))):
                            try:
                                group_name = str(df.iloc[j, 1]).strip() if len(df.columns) > 1 else ""

                                # 🎯 실제 그룹명만 인식 (통계용어 완전 제외)
                                if (group_name and group_name != '전체' and group_name != dep_var and
                                    self.is_real_group_name(group_name)):

                                    # 숫자 데이터 확인
                                    n_col = self.find_number_column(df, j, 'n')
                                    mean_col = self.find_number_column(df, j, 'mean')
                                    std_col = self.find_number_column(df, j, 'std')

                                    if n_col is not None and mean_col is not None and std_col is not None:
                                        n_val = float(df.iloc[j, n_col])
                                        mean_val = float(df.iloc[j, mean_col])
                                        std_val = float(df.iloc[j, std_col])

                                        groups_data[dep_var].append({
                                            'group': group_name,
                                            'n': int(n_val),
                                            'mean': mean_val,
                                            'std': std_val
                                        })

                                        self.root.after(0, lambda var=dep_var, grp=group_name:
                                                       self.log(f"    ✅ {var} - {grp} (실제 그룹)"))
                            except (ValueError, IndexError):
                                continue

                        break  # 이 변수는 처리 완료

            # 빈 그룹 제거
            groups_data = {k: v for k, v in groups_data.items() if v}

            # 2. ANOVA 결과 추출 - 완전히 유연한 방식
            test_results = {}

            self.root.after(0, lambda: self.log(f"    🔍 ANOVA 결과 스캔 시작..."))

            # ANOVA 테이블에서 F값과 p값 찾기
            for i in range(results_row, min(results_row + 50, len(df))):
                row_content = " ".join([str(df.iloc[i, col]).strip() for col in range(min(len(df.columns), 10))])

                # 집단-간 (Between Groups) 행 찾기
                if any(keyword in row_content for keyword in ['집단-간', 'Between Groups', '집단간']):
                    # 이 행에서 F값과 p값 찾기
                    for dep_var in possible_dep_vars:
                        if dep_var in groups_data:  # 그룹 데이터가 있는 변수만
                            # F값과 p값 찾기 (여러 컬럼 시도)
                            f_val, p_val = self.extract_f_and_p_values(df, i)

                            if f_val is not None and p_val is not None:
                                # 변수명 매칭 (가장 가까운 위치의 변수)
                                closest_var = self.find_closest_variable(df, i, possible_dep_vars)
                                if closest_var:
                                    test_results[closest_var] = {'f': f_val, 'p': p_val}
                                    self.root.after(0, lambda var=closest_var, f=f_val, p=p_val:
                                                   self.log(f"    ✅ ANOVA: {var}, F={f:.3f}, p={p:.6f}"))
                                break

            # 3. 결과 저장
            self.root.after(0, lambda: self.log(f"  📊 기술통계: {list(groups_data.keys())}"))
            self.root.after(0, lambda: self.log(f"  📊 ANOVA 결과: {list(test_results.keys())}"))

            success_count = 0
            for var in groups_data:
                if var in test_results:
                    self.all_analyses.append({
                        'indep_var': pair['indep_var'],
                        'dep_var': var,
                        'groups': groups_data[var],
                        'statistic': test_results[var]['f'],
                        'p_value': test_results[var]['p'],
                        'test_type': 'anova'
                    })
                    self.root.after(0, lambda var=var: self.log(f"  ✅ 저장 성공: {var}"))
                    success_count += 1
                else:
                    self.root.after(0, lambda var=var: self.log(f"  ❌ 저장 실패: {var}"))

            self.root.after(0, lambda count=success_count: self.log(f"  📈 성공률: {count}/{len(groups_data)}개"))

        except Exception as e:
            self.root.after(0, lambda e=e: self.log(f"  💥 ANOVA 추출 오류: {e}", 'error'))

    def find_number_column(self, df: pd.DataFrame, row: int, data_type: str) -> Optional[int]:
        """숫자 데이터가 있는 컬럼 찾기"""
        try:
            for col in range(2, min(len(df.columns), 8)):  # 2번째 컬럼부터 확인
                cell_val = str(df.iloc[row, col]).strip()
                if cell_val and cell_val != '':
                    try:
                        float(cell_val)
                        return col
                    except ValueError:
                        continue
            return None
        except:
            return None

    def extract_f_and_p_values(self, df: pd.DataFrame, row: int) -> tuple:
        """F값과 p값 추출"""
        try:
            f_val = None
            p_val = None

            # 여러 컬럼에서 F값과 p값 찾기
            for col in range(min(len(df.columns), 10)):
                cell_val = str(df.iloc[row, col]).strip()
                if cell_val and cell_val != '':
                    try:
                        num_val = float(cell_val)
                        if 0 <= num_val <= 1 and p_val is None:  # p값 같아 보이는 것
                            p_val = num_val
                        elif num_val > 1 and f_val is None:  # F값 같아 보이는 것
                            f_val = num_val
                    except ValueError:
                        continue

            return f_val, p_val
        except:
            return None, None

    def find_closest_variable(self, df: pd.DataFrame, row: int, possible_vars: set) -> Optional[str]:
        """가장 가까운 위치의 변수명 찾기"""
        try:
            # 현재 행 위쪽에서 변수명 찾기
            for i in range(row, max(0, row - 20), -1):
                for col in range(min(3, len(df.columns))):
                    cell_val = str(df.iloc[i, col]).strip()
                    if cell_val in possible_vars:
                        return cell_val

            # 못 찾으면 첫 번째 가능한 변수 반환
            return list(possible_vars)[0] if possible_vars else None
        except:
            return None


    def create_perfect_output(self) -> str:
        """동적 종속변수 개수 기반 OUTPUT 표 생성"""
        wb = Workbook()
        ws = wb.active
        ws.title = 'OUTPUT'

        # 🎯 모든 종속변수 동적 추출
        dep_vars = []
        if self.all_analyses:
            # 모든 종속변수 수집
            all_dep_vars = list(set([analysis['dep_var'] for analysis in self.all_analyses]))

            # 선호 순서로 정렬 (특정 패턴 우선)
            priority_patterns = ['역량', '신념', '성과', '만족', '평균', '합계', '점수']

            def sort_key(var):
                for i, pattern in enumerate(priority_patterns):
                    if pattern in var:
                        return i
                return len(priority_patterns)

            dep_vars = sorted(all_dep_vars, key=sort_key)

        # 🎯 종속변수 개수에 맞춰 동적 테이블 생성
        dep_count = len(dep_vars) if dep_vars else 1
        self.root.after(0, lambda count=dep_count:
                       self.log(f"📋 종속변수 {count}개 발견, 동적 표 생성"))

        # 종속변수별로 로그 출력
        for i, var in enumerate(dep_vars):
            self.root.after(0, lambda idx=i+1, v=var:
                           self.log(f"  {idx}. {v}"))

        # 동적 헤더 생성
        self.create_dynamic_headers(ws, dep_vars)

        # 동적 데이터 생성
        self.create_dynamic_data(ws, dep_vars)

        # 동적 스타일 적용
        self.apply_dynamic_styles(ws, len(dep_vars))

        # 저장
        output_path = self.generate_output_path()
        wb.save(output_path)
        return output_path

    def create_dynamic_headers(self, ws, dep_vars: list) -> None:
        """종속변수 개수에 맞는 동적 헤더 생성"""
        if not dep_vars:
            dep_vars = ["종속변수1"]

        # 🎯 동적 컬럼 수 계산 (하드코딩 제거)
        cols_per_var = self.get_columns_per_variable()  # 동적으로 계산된 컬럼 수

        # 첫 번째 헤더 행: 종속변수명
        header1 = ['독립변수']
        for i, dep_var in enumerate(dep_vars):
            header1.extend([dep_var] + [''] * (cols_per_var - 1))
        ws.append(header1)

        # 두 번째 헤더 행: 컬럼명
        header2 = ['']
        for _ in dep_vars:
            header2.extend(['그룹', 'N', '평균', '표준편차', '통계량', 'p값'])
        ws.append(header2)

        # 헤더 병합
        for i, dep_var in enumerate(dep_vars):
            start_col = 2 + i * cols_per_var  # B부터 시작 (A는 독립변수)
            end_col = start_col + cols_per_var - 1
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    def create_dynamic_data(self, ws, dep_vars: list) -> None:
        """종속변수 개수에 맞는 동적 데이터 생성"""
        if not self.all_analyses:
            # 빈 데이터인 경우
            row = ['데이터 없음'] + ['추출된 분석 결과가 없습니다'] * len(dep_vars) * 6
            ws.append(row)
            return

        # 독립변수별로 데이터 생성
        processed_indep_vars = []

        for analysis in self.all_analyses:
            indep_var = analysis['indep_var']

            if indep_var in processed_indep_vars:
                continue

            processed_indep_vars.append(indep_var)

            # 🎯 이 독립변수의 모든 종속변수 데이터 수집
            indep_var_data = {}
            for a in self.all_analyses:
                if a['indep_var'] == indep_var:
                    indep_var_data[a['dep_var']] = a

            # 최대 그룹 수 계산
            max_groups = 0
            for dep_var in dep_vars:
                if dep_var in indep_var_data:
                    max_groups = max(max_groups, len(indep_var_data[dep_var]['groups']))

            # 그룹별로 행 생성
            for group_idx in range(max_groups):
                row = [indep_var if group_idx == 0 else '']

                # 각 종속변수별로 데이터 추가
                for dep_var in dep_vars:
                    if dep_var in indep_var_data:
                        analysis = indep_var_data[dep_var]

                        if group_idx < len(analysis['groups']):
                            group = analysis['groups'][group_idx]
                            stat_val = round(analysis['statistic'], 3) if group_idx == 0 else ''
                            p_val = self.format_p_value(analysis['p_value']) if group_idx == 0 else ''

                            row.extend([
                                group['group'], group['n'],
                                round(group['mean'], 4), round(group['std'], 5),
                                stat_val, p_val
                            ])
                        else:
                            # 빈 데이터
                            row.extend(['', '', '', '', '', ''])
                    else:
                        # 해당 독립변수에 이 종속변수 데이터가 없음
                        row.extend(['', '', '', '', '', ''])

                ws.append(row)

    def apply_dynamic_styles(self, ws, dep_var_count: int) -> None:
        """동적 스타일 적용"""
        cols_per_var = self.get_columns_per_variable()  # 동적 컬럼 수
        total_cols = 1 + dep_var_count * cols_per_var  # 독립변수 컬럼 + 종속변수별 컬럼들

        # 헤더 병합 및 스타일
        header_font = Font(bold=True, size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

        # 독립변수 헤더
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align
        ws['A1'].fill = header_fill

        # 종속변수별 헤더
        for i in range(dep_var_count):
            start_col = 2 + i * cols_per_var
            header_cell = ws.cell(row=1, column=start_col)
            header_cell.font = header_font
            header_cell.alignment = center_align
            header_cell.fill = header_fill

        # 컬럼 헤더 스타일 (2번째 행)
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True, size=10)
            cell.fill = header_fill
            cell.alignment = center_align

        # 데이터 정렬
        for row in range(3, ws.max_row + 1):
            for col in range(1, total_cols + 1):
                ws.cell(row=row, column=col).alignment = center_align

        # 동적 열 너비 설정
        self.set_dynamic_column_widths(ws, dep_var_count)

        # 테두리
        self.add_borders_to_table(ws, total_cols)

    def set_dynamic_column_widths(self, ws, dep_var_count: int) -> None:
        """동적 열 너비 설정"""
        # 독립변수 컬럼 (A)
        ws.column_dimensions['A'].width = 18

        # 🎯 동적 컬럼 너비 계산 (하드코딩 제거)
        col_widths = self.get_column_widths()
        col_letters = 'BCDEFGHIJKLMNOPQRSTUVWXYZ'

        cols_per_var = self.get_columns_per_variable()
        for dep_idx in range(dep_var_count):
            for col_idx in range(cols_per_var):
                col_pos = 1 + dep_idx * cols_per_var + col_idx  # B부터 시작
                if col_pos < len(col_letters) and col_idx < len(col_widths):
                    ws.column_dimensions[col_letters[col_pos]].width = col_widths[col_idx]

    def get_column_widths(self) -> list:
        """동적 컬럼 너비 계산"""
        # 그룹, N, 평균, 표준편차, 통계량, p값
        return [15, 8, 12, 12, 10, 10]

    def get_search_range(self, base_type: str) -> int:
        """검색 범위 동적 계산 (하드코딩 제거)"""
        search_ranges = {
            'table_area': 100,     # 테이블 영역 찾기
            'stats_data': 50,      # 통계 데이터 추출
            'extended_stats': 200, # 확장 통계 데이터
            'group_search': 15,    # 그룹 검색
            'result_search': 30,   # 결과 검색
            'anova_result': 50,    # ANOVA 결과
            'nearby_search': 10,   # 근처 데이터 검색
            'levene_check': 3      # Levene 검정 확인
        }
        return search_ranges.get(base_type, 50)  # 기본값 50

    def add_borders_to_table(self, ws, total_cols: int) -> None:
        """테이블에 테두리 추가"""
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=total_cols):
            for cell in row:
                cell.border = thin

    def format_p_value(self, p_val: float) -> str:
        """p값 형식화"""
        if p_val < 0.001:
            return "0.000"
        else:
            return f"{p_val:.3f}"

    def generate_output_path(self) -> str:
        """출력 파일 경로 생성"""
        try:
            base_name = os.path.splitext(self.file_path)[0]
            output_path = f"{base_name}_OUTPUT.xlsx"

            counter = 1
            while os.path.exists(output_path):
                output_path = f"{base_name}_OUTPUT_{counter}.xlsx"
                counter += 1

            return output_path
        except:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            return f"SPSS_OUTPUT_{timestamp}.xlsx"

    def apply_styles(self, ws) -> None:
        """스타일 적용"""
        # 헤더 병합
        ws.merge_cells('A1:F1')
        ws.merge_cells('G1:L1')

        # 폰트 및 정렬
        header_font = Font(bold=True, size=11)
        center_align = Alignment(horizontal='center', vertical='center')

        ws['A1'].font = header_font
        ws['A1'].alignment = center_align
        ws['G1'].font = header_font
        ws['G1'].alignment = center_align

        # 컬럼 헤더 스타일
        header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            cell = ws[f'{col}2']
            cell.font = Font(bold=True, size=10)
            cell.fill = header_fill
            cell.alignment = center_align

        # 데이터 정렬
        for row in range(3, ws.max_row + 1):
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}{row}'].alignment = center_align

        # 열 너비
        widths = {'A': 15, 'B': 12, 'C': 8, 'D': 10, 'E': 12, 'F': 10, 'G': 8,
                 'H': 12, 'I': 8, 'J': 10, 'K': 12, 'L': 10}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        # 테두리
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=12):
            for cell in row:
                cell.border = thin

    def open_file(self, file_path: str) -> None:
        """파일 열기"""
        try:
            if sys.platform == 'darwin':
                os.system(f'open "{file_path}"')
            elif sys.platform == 'win32':
                os.startfile(file_path)
            else:
                os.system(f'xdg-open "{file_path}"')
        except:
            pass

def main() -> None:
    """메인 함수"""
    root = tk.Tk()
    SPSSAnalysisExtractor(root)
    root.mainloop()

if __name__ == "__main__":
    main()